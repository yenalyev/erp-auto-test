#!/usr/bin/env python3
"""
Generate TCM import XLSX — gaps vs export tcm-ERP__________-project-20260702.xlsx.

96 automation @TestCaseId missing from TCM (333 cases in export vs 370 in code).
Focus: manual QA — clear title, business goal in description, step-by-step actions.
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from tcm_import_common import (
    ROLE_ADMIN,
    ROLE_OWNER,
    Case,
    Step,
    write_xlsx_with_features,
)

OUTPUT = Path(__file__).resolve().parent.parent / "docs" / "tcm-import-gap-20260702.xlsx"
TCM_EXPORT = Path(r"C:\Users\gigam\Downloads\tcm-ERP__________-project-20260702.xlsx")

# --- Feature ids (existing in TCM unless noted NEW) ---
FEAT_ACC = "REQ-ACC"
FEAT_STR = "REQ-STR"
FEAT_STR_DEACT = "REQ-STR"  # AC-03 — деактивація локацій
FEAT_EDIT_REL = "REQ-EDIT_REL"
FEAT_REGION_LOC = "REQ-REGION-001"
FEAT_WMS_SEND = "REQ-WMS-001"
FEAT_WMS_INV = "REQ-WMS-003"
FEAT_WMS_JOURNAL = "REQ-WMS-006"
FEAT_WMS_INVOICE = "REQ-WMS-005"
FEAT_WMS_STOCK = "REQ-WMS-007"
FEAT_MFG_TM_ARCH = "REQ-MFG-001-04"
FEAT_GP = "REQ-GLOBAL-PLAN"
FEAT_PLAN = "REQ-MAN-PLN"
FEAT_EQU = "REQ-EQU-001"
FEAT_RES_RBAC = "REQ-RES-003"
FEAT_CREW_ROOT = "REQ-CREW"          # NEW
FEAT_CREW_REG = "REQ-CREW-001"       # NEW
FEAT_CREW_ISSUE = "REQ-CREW-002"     # NEW
FEAT_CREW_INV = "REQ-CREW-003"       # NEW
FEAT_RVW = "REQ-RVW"                 # NEW

ROLE_ACC = "Accounter"
ROLE_CM = "Crew-Manager"
ROLE_RVW = "ResourceViewer"

PRE_ADMIN = "@Admin залогінений. Середовище dev/staging."
PRE_OWNER1 = (
    "@Owner 1 (alkatras) залогінений. У шапці SPA обрана конкретна локація Owner 1 "
    "(не «Всі локації»)."
)
PRE_OWNER2 = (
    "@Owner 2 (bar) залогінений. Підрозділ accessMode=REGIONS. "
    "У шапці обрана локація Owner 2."
)
PRE_ACC = "@Accounter (accountant) залогінений."
PRE_CREW = (
    f"{PRE_ADMIN} Підготовлена область accessMode=CREWS: UNIT-отримувач, CREW-елемент, "
    "member-підрозділ з hasCrews=true. На member-складі stock ресурсу ≥ 50 од."
)
PRE_CM = f"{PRE_CREW} Сесія Crew-Manager (argument) з правом inventory-list::{{crew}}::read."


def mk(
    test_id: str,
    feature_id: str,
    ac_id: str,
    title: str,
    goal: str,
    *,
    priority: str = "HIGH",
    severity: str = "MAJOR",
    preconditions: str = "",
    expected_result: str = "",
    tags: str = "",
    role_name: str = ROLE_ADMIN,
    layer: str | None = None,
    steps: list[tuple[str, str]],
    cross: list[str] | None = None,
    test_data: str = "",
) -> Case:
    desc_parts = [f"Мета: {goal}"]
    if test_data:
        desc_parts.append(f"Тестові дані: {test_data}")
    description = "\n".join(desc_parts)
    return Case(
        test_id=test_id,
        feature_id=feature_id,
        ac_id=ac_id,
        title=title,
        description=description,
        priority=priority,
        severity=severity,
        preconditions=preconditions,
        expected_result=expected_result or (steps[-1][1] if steps else ""),
        tags=tags,
        role_name=role_name,
        automation_layer=None,
        automation_test_id=None,
        cross_features=cross or [],
        steps=[Step(i + 1, a, e) for i, (a, e) in enumerate(steps)],
    )


# --------------------------------------------------------------------------- #
# Domain case builders
# --------------------------------------------------------------------------- #

def accountant_cases() -> list[Case]:
    return [
        mk(
            "TC-ACC-API-001", FEAT_ACC, "AC-01",
            "API — селектор my-units для бухгалтера без UNIT-локацій",
            "Бухгалтер у робочому просторі бачить лише операційні локації (STORAGE/PRODUCTION), "
            "а не підрозділи type=UNIT — це обмежує доступ до внутрішньої ієрархії.",
            priority="CRITICAL", preconditions=PRE_ACC, tags="accountant,api,my-units",
            role_name=ROLE_ACC,
            steps=[
                ("Увійти як @Accounter (API-сесія або DevTools → Application → cookies)", "Сесія активна"),
                ("Виконати GET /api/v1/storages/names/my-units", "HTTP 200; список не порожній"),
                ("Перевірити кожен елемент: type ≠ UNIT", "Жодного запису з type=UNIT у відповіді"),
            ],
        ),
        mk(
            "TC-ACC-API-002", FEAT_ACC, "AC-02",
            "API — журнал переміщень бухгалтера: лише STORAGE→UNIT",
            "Бізнес-контракт логістики для бухгалтера: у журналі видно лише видачі зі складу на підрозділ; "
            "внутрішні рухи UNIT→CREW та UNIT→UNIT приховані.",
            priority="CRITICAL",
            preconditions=f"{PRE_ADMIN} Створено 3 переміщення: UNIT→CREW, UNIT→UNIT, STORAGE→UNIT (по 5 од. ресурсу).",
            test_data="OWNER_1 sender; crew + другий UNIT + owner2 storage як recipients; resource acc-rel-*.",
            tags="accountant,relocations,api", role_name=ROLE_ACC,
            steps=[
                ("Під OWNER_1 створити видачу UNIT → CREW (AUTO_FINISHED)", "relocationId #1"),
                ("Під OWNER_1 створити видачу UNIT → UNIT", "relocationId #2"),
                ("Під OWNER_1 створити видачу STORAGE → UNIT (owner1 → owner2)", "relocationId #3"),
                ("GET /api/v1/relocations (sent history) як ACCOUNTANT", "HTTP 200"),
                ("Перевірити id у content", "Є #3; немає #1 і #2"),
            ],
        ),
        mk(
            "TC-UI-ACC-001", FEAT_ACC, "AC-01",
            "UI — кабінет бухгалтера після логіну (layout smoke)",
            "Після входу бухгалтер потрапляє на /production і бачить обмежений набір розділів sidebar, "
            "селектор локацій та фільтри журналу — без адмінських словників.",
            priority="CRITICAL", preconditions="Облікові дані @Accounter на dev/staging.",
            tags="accountant,ui,smoke", role_name=ROLE_ACC, layer="UI",
            steps=[
                ("Відкрити SPA → редірект на Keycloak", "Форма логіну видима"),
                ("Увійти валідними credentials accountant", "Редірект на /production"),
                ("Перевірити sidebar: Виробництво, Несерійне, Залишки, Обладнання, Логістика, Експорт", "Усі розділи видимі"),
                ("Перевірити словники: Техкарти, Словник ресурсів, Ціни", "Пункти видимі"),
                ("Перевірити селектор «Робочий простір» — обрана перша доступна локація", "selected = перша в списку"),
                ("Перевірити фільтри журналу: Продукт, Категорія, Тип робіт, З, По", "Фільтри відрендерені"),
            ],
        ),
        mk(
            "TC-UI-ACC-002", FEAT_ACC, "AC-02",
            "UI — вихід бухгалтера з системи",
            "Користувач може завершити сесію через меню профілю в sidebar.",
            priority="CRITICAL", preconditions=PRE_ACC, tags="accountant,ui,auth",
            role_name=ROLE_ACC, layer="UI",
            steps=[
                ("Увійти як accountant", "SPA завантажено"),
                ("Відкрити меню користувача в sidebar → «Вийти»", "Сесія завершена"),
                ("Перевірити URL / форму", "Повернення на Keycloak login"),
            ],
        ),
        mk(
            "TC-UI-ACC-003", FEAT_ACC, "AC-01",
            "UI — робочий простір бухгалтера без UNIT у my-units",
            "UI-селектор локацій відповідає API: бухгалтер не обирає підрозділи type=UNIT.",
            priority="CRITICAL", preconditions=PRE_ACC, tags="accountant,ui,my-units",
            role_name=ROLE_ACC, layer="UI",
            steps=[
                ("Увійти як accountant", "Сторінка завантажена"),
                ("DevTools → Network: знайти GET /storages/names/my-units", "HTTP 200"),
                ("Перевірити JSON: жоден type ≠ UNIT", "UNIT відсутні"),
                ("Відкрити dropdown «Робочий простір» у sidebar", "Список локацій не порожній"),
            ],
        ),
        mk(
            "TC-UI-ACC-004", FEAT_ACC, "AC-02",
            "UI — логістика бухгалтера приховує UNIT→CREW",
            "На вкладці «Видано» розділу Логістика не відображаються видачі на екіпажі.",
            priority="HIGH",
            preconditions=f"{PRE_ADMIN} API: створено UNIT→CREW видачу (crew acc-ui-crew-*).",
            test_data="Після fixture — логін accountant, обрана локація owner1 storage.",
            tags="accountant,ui,relocations", role_name=ROLE_ACC, layer="UI",
            steps=[
                ("Увійти як accountant, обрати локацію owner1", "Сесія + workspace"),
                ("Відкрити /relocations → Логістика → вкладка «Видано»", "Журнал завантажено"),
                ("Знайти рядок з іменем crew з fixture", "Рядок відсутній у таблиці"),
            ],
        ),
    ]


def storage_crud_cases() -> list[Case]:
    pre = PRE_ADMIN
    return [
        mk("TC-STR-004", FEAT_STR, "AC-01", "Локація — валідація обов'язкового поля name",
           "Система відхиляє створення локації без назви — захист довідника від «порожніх» записів.",
           preconditions=pre, tags="storage,validation,negative",
           test_data="POST /api/v1/storages: name=null, name='', name='   ' (3 варіанти).",
           steps=[
               ("POST /storages з name=null", "HTTP 400"),
               ("POST /storages з name=''", "HTTP 400"),
               ("POST /storages з name лише пробіли", "HTTP 400"),
           ]),
        mk("TC-STR-005", FEAT_STR, "AC-01", "Локація — заборона INTERNAL→EXTERNAL при PUT",
           "Після підтвердження внутрішньої локації (INTERNAL) не можна повернути її у зовнішній контур — "
           "захист цілісності обліку.",
           preconditions=pre, tags="storage,validation,relation",
           test_data="Існуюча INTERNAL child STORAGE; PUT relation=EXTERNAL.",
           steps=[
               ("Створити INTERNAL child STORAGE", "relation=INTERNAL"),
               ("PUT /storages/{id} з relation=EXTERNAL", "HTTP 400; relation лишається INTERNAL"),
           ]),
        mk("TC-STR-006", FEAT_STR_DEACT, "AC-03", "Локація — архівація (деактивація)",
           "Архівована локація зникає з активних селекторів, але залишається в архіві для аудиту.",
           priority="CRITICAL", preconditions=pre, tags="storage,deactivation",
           steps=[
               ("Створити ізольовану локацію", "id отримано"),
               ("DELETE /storages/{id} (архівація)", "HTTP 200; active=false"),
               ("GET /storages/names?isActive=true — пошук за ім'ям", "Локація відсутня"),
               ("GET /storages/names?isActive=false", "Локація присутня"),
           ]),
        mk("TC-STR-007", FEAT_STR_DEACT, "AC-03", "Локація — розархівація (unarchive)",
           "Адмін може відновити архівну локацію — вона знову з'являється в активних списках.",
           priority="CRITICAL", preconditions=pre, tags="storage,deactivation",
           steps=[
               ("Створити і заархівувати локацію", "active=false"),
               ("PUT /storages/unarchive/{id}", "HTTP 200; active=true"),
               ("GET /storages/names?isActive=true", "Локація знову в списку"),
           ]),
        mk("TC-STR-008", FEAT_STR_DEACT, "AC-03", "Локація — повторне використання назви після архівації",
           "Унікальність name лише серед active; назва архівної локації доступна для нового POST.",
           preconditions=pre, tags="storage,deactivation",
           steps=[
               ("Створити локацію «Test-X», заархівувати", "Архів"),
               ("POST нова локація з name=«Test-X»", "HTTP 200; нова active локація"),
           ]),
        mk("TC-STR-009", FEAT_STR, "AC-01", "Локація — POST поля довжиною 255 символів",
           "Граничні значення name/alias/identifierNumber/nameForInvoices приймаються (макс. довжина).",
           preconditions=pre, tags="storage,validation",
           test_data="name, alias, identifierNumber, nameForInvoices — рівно 255 символів.",
           steps=[
               ("POST /storages з усіма текстовими полями = 255 символів", "HTTP 200"),
               ("GET /storages/{id}", "Поля збережені повністю"),
           ]),
        mk("TC-STR-010", FEAT_STR, "AC-01", "Локація — POST відхиляє 256 символів",
           "Перевищення ліміту довжини полів → HTTP 400 (не 500).",
           preconditions=pre, tags="storage,validation,negative",
           test_data="name або alias = 256 символів.",
           steps=[("POST /storages з name=256 символів", "HTTP 400")]),
        mk("TC-STR-011", FEAT_STR, "AC-01", "Локація — PUT приймає 255 символів",
           "Оновлення існуючої локації з граничною довжиною полів.",
           preconditions=pre, tags="storage,validation",
           steps=[
               ("Створити локацію", "id відомий"),
               ("PUT з name/alias = 255 символів", "HTTP 200"),
           ]),
        mk("TC-STR-012", FEAT_STR, "AC-01", "Локація — PUT відхиляє 256 символів",
           "Валідація довжини при оновленні.",
           preconditions=pre, tags="storage,validation,negative",
           steps=[
               ("Створити локацію", "id відомий"),
               ("PUT з name=256 символів", "HTTP 400"),
           ]),
        mk("TC-STR-013", FEAT_STR, "AC-01", "Локація — створення EXTERNAL relation",
           "За замовчуванням нова дочірня STORAGE може мати relation=EXTERNAL (зовнішній контур).",
           priority="CRITICAL", preconditions=pre, tags="storage,relation",
           test_data="type=STORAGE, parentId=unit, accessMode=FULL_ACCESS, relation=EXTERNAL.",
           steps=[
               ("POST child STORAGE з relation=EXTERNAL", "HTTP 200"),
               ("GET /storages/{id}", "relation=EXTERNAL у response"),
           ]),
        mk("TC-STR-014", FEAT_STR, "AC-01", "Локація — перехід EXTERNAL→INTERNAL дозволений",
           "Односторонній перехід у внутрішній контур (імітація confirm у UI).",
           preconditions=pre, tags="storage,relation",
           steps=[
               ("Створити EXTERNAL STORAGE", "relation=EXTERNAL"),
               ("PUT relation=INTERNAL", "HTTP 200; GET підтверджує INTERNAL"),
           ]),
        mk("TC-STR-015", FEAT_STR, "AC-01", "Локація — GET фільтр relation=INTERNAL",
           "Query ?relation=INTERNAL фільтрує за відношенням, не за type.",
           preconditions=pre, tags="storage,relation",
           test_data="Пара child: INTERNAL (int-filter-) та EXTERNAL (ext-filter-) під одним parent.",
           steps=[
               ("Створити INTERNAL і EXTERNAL child під одним parent", "2 id"),
               ("GET /storages?relation=INTERNAL", "Є INTERNAL id; немає EXTERNAL id"),
           ]),
        mk("TC-STR-016", FEAT_STR, "AC-01", "API — /names?relation=EXTERNAL для селектора «Звідки»",
           "Контракт dropdown «Звідки» на /equipment: лише EXTERNAL локації.",
           priority="CRITICAL", preconditions=pre, tags="storage,equipment,relation",
           steps=[
               ("Створити INTERNAL і EXTERNAL child STORAGE", "2 id"),
               ("GET /storages/names?isActive=true&relation=EXTERNAL", "EXTERNAL є; INTERNAL немає"),
           ]),
        mk("TC-STR-017", FEAT_STR, "AC-01", "Локація — GET ?relation=EXTERNAL + schema",
           "Сторінковий список EXTERNAL локацій проходить JSON schema.",
           preconditions=pre, tags="storage,relation",
           steps=[
               ("GET /storages?relation=EXTERNAL&size=500", "HTTP 200; schema valid"),
               ("Перевірити кожен елемент content", "Усі relation=EXTERNAL"),
           ]),
        mk("TC-STR-018", FEAT_STR, "AC-01", "Локація — relation не залежить від UnitType",
           "INTERNAL/EXTERNAL семантика однакова для STORAGE, UNIT, PRODUCTION.",
           priority="CRITICAL", preconditions=pre, tags="storage,relation",
           test_data="Для кожного type ∈ {STORAGE, UNIT, PRODUCTION}: пара INTERNAL + EXTERNAL child.",
           steps=[
               ("Для type=STORAGE: створити INTERNAL + EXTERNAL", "2 id"),
               ("Перевірити ?relation=INTERNAL та ?relation=EXTERNAL", "Кожен id лише у своєму фільтрі"),
               ("Повторити для type=UNIT та PRODUCTION", "Та сама логіка для обох"),
           ]),
    ]


def storage_relation_inventory_cases() -> list[Case]:
    pre = PRE_ADMIN
    return [
        mk("TC-INV-REL-001", FEAT_STR, "AC-01", "INTERNAL receive — залишок збільшується",
           "Прийом на внутрішню локацію (INTERNAL) збільшує stock отримувача.",
           preconditions=pre, tags="storage,relation,stock",
           steps=[
               ("Зафіксувати stock recipient (INTERNAL)", "before"),
               ("POST receive SUPPLIER → INTERNAL storage", "HTTP 200; FINISHED"),
               ("Перевірити stock", "after = before + amount"),
           ]),
        mk("TC-INV-REL-002", FEAT_STR, "AC-01", "EXTERNAL receive — залишок без змін (no-op)",
           "Зовнішній контур не веде управлінський облік залишків через relocation receive.",
           preconditions=pre, tags="storage,relation,stock",
           steps=[
               ("Зафіксувати stock EXTERNAL storage", "before"),
               ("POST receive на EXTERNAL", "HTTP 200"),
               ("Перевірити stock", "after = before"),
           ]),
        mk("TC-INV-REL-003", FEAT_STR, "AC-01", "INTERNAL send+resolve — залишок на recipient",
           "Внутрішнє переміщення зараховує товар на склад отримувача після resolve.",
           preconditions=pre, tags="storage,relation,stock",
           steps=[
               ("POST send INTERNAL→INTERNAL (CREATED)", "relocationId"),
               ("POST resolve FINISHED", "state=FINISHED"),
               ("Перевірити stock recipient", "+amount"),
           ]),
        mk("TC-INV-REL-004", FEAT_WMS_INV, "AC-01", "EXTERNAL — inventory PUT змінює залишок (WMS шлях)",
           "На EXTERNAL локації залишки змінюються через інвентаризацію, не через relocation.",
           preconditions=pre, tags="storage,relation,inventory",
           steps=[
               ("Відкрити inventory session на EXTERNAL", "session open"),
               ("PUT /inventory з target amount=25", "HTTP 200"),
               ("Перевірити stock", "amount=25"),
           ]),
        mk("TC-INV-REL-005", FEAT_STR, "AC-01", "EXTERNAL receive — no-op для STORAGE/UNIT/PRODUCTION",
           "Receive на EXTERNAL не змінює stock незалежно від type локації.",
           preconditions=pre, tags="storage,relation,stock",
           test_data="EXTERNAL child для type ∈ {STORAGE, UNIT, PRODUCTION}.",
           steps=[
               ("Для кожного type: receive на EXTERNAL child", "HTTP 200"),
               ("Перевірити stock після кожного", "stock без змін"),
           ]),
        mk("TC-REL-REL-001", FEAT_EDIT_REL, "AC-02", "Видача на EXTERNAL — AUTO_FINISHED, stock −",
           "Видача на зовнішній контур списує з відправника, але не зараховує на EXTERNAL recipient.",
           preconditions=PRE_OWNER1, tags="relocations,relation,stock", role_name=ROLE_OWNER,
           test_data="sender INTERNAL; EXTERNAL child recipient; amount=4.",
           steps=[
               ("Зафіксувати stock sender і EXTERNAL recipient", "snapshots"),
               ("POST send sender → EXTERNAL", "state=AUTO_FINISHED"),
               ("Перевірити stock", "sender −4; EXTERNAL без змін"),
           ]),
        mk("TC-REL-REL-002", FEAT_EDIT_REL, "AC-02", "INTERNAL→INTERNAL resolve — stock + на recipient",
           "Класичне внутрішнє переміщення: після підтвердження отримувачем залишок збільшується.",
           preconditions=PRE_OWNER1, tags="relocations,relation,stock", role_name=ROLE_OWNER,
           steps=[
               ("POST send INTERNAL→INTERNAL", "CREATED"),
               ("POST resolve FINISHED", "FINISHED"),
               ("Перевірити stock recipient", "+amount"),
           ]),
    ]


def crew_region_cases() -> list[Case]:
    api = "crew-regions,api"
    return [
        mk("TC-STR-CREW-001", FEAT_CREW_REG, "AC-01", "Створення області accessMode=CREWS",
           "Адмін налаштовує область видимості екіпажів: named-набір UNIT+CREW для видачі ресурсів на екіпажі.",
           priority="CRITICAL", preconditions=PRE_ADMIN, tags=api,
           test_data="recipient UNIT crew-reg-rec-*; accessMode=CREWS.",
           steps=[
               ("Створити UNIT recipient", "storageId"),
               ("POST /storages/regions accessMode=CREWS", "HTTP 200; accessMode=CREWS"),
           ]),
        mk("TC-STR-CREW-002", FEAT_CREW_REG, "AC-02", "Locations області CREWS містять UNIT",
           "У області мають бути UNIT-вузли для каскадного пошуку екіпажів у формі видачі.",
           preconditions=PRE_CREW, tags=api,
           steps=[
               ("Підготувати CrewRegionScenario", "regionId, unitId"),
               ("GET /storages/regions/{id}/locations", "HTTP 200; містить unit.id"),
           ]),
        mk("TC-STR-CREW-003", FEAT_CREW_REG, "AC-03", "Members області CREWS — підрозділ-споживач",
           "Member storage отримує hasCrews=true і доступ до видачі на екіпажі.",
           preconditions=PRE_CREW, tags=api,
           steps=[
               ("Додати memberStorage до region members", "HTTP 200"),
               ("GET /storages/regions/{id}/members", "memberStorageId у списку"),
           ]),
        mk("TC-STR-CREW-004", FEAT_CREW_REG, "AC-04", "GET області CREWS за id",
           "Деталі області: accessMode=CREWS, recipientStorage = UNIT.",
           preconditions=PRE_CREW, tags=api,
           steps=[("GET /storages/regions/{id}", "accessMode=CREWS; recipient.id = unit.id")]),
        mk("TC-STR-CREW-005", FEAT_CREW_REG, "AC-05", "hasCrews=true для member у області CREWS",
           "Creation-options показує наявність екіпажів для member storage в області.",
           preconditions=PRE_CREW, tags=api,
           steps=[
               ("GET /relocations/creation-options?storageId={member}", "hasCrews=true"),
               ("Для storage поза областю", "hasCrews=false"),
           ]),
        mk("TC-STR-CREW-006", FEAT_CREW_REG, "AC-06", "Область CREWS без members — hasCrews=false",
           "Без members жоден підрозділ не бачить кнопку видачі на екіпаж.",
           preconditions=PRE_ADMIN, tags=api,
           steps=[
               ("Створити CREWS region з locations, members=[]", "regionId"),
               ("GET creation-options для довільного storage", "hasCrews=false"),
           ]),
        mk("TC-STR-CREW-011", FEAT_CREW_REG, "AC-07", "GET /names/crew-units — ієрархія UNIT",
           "API повертає дерево UNIT для каскадного combobox «Підрозділ» у формі видачі.",
           preconditions=PRE_CREW, tags=api,
           steps=[("GET /storages/names/crew-units", "HTTP 200; містить очікувані UNIT nodes")]),
        mk("TC-STR-CREW-012", FEAT_CREW_REG, "AC-08", "GET /names/crews?parentId= — екіпажі під UNIT",
           "Після вибору UNIT список екіпажів фільтрується за parentId.",
           preconditions=PRE_CREW, tags=api,
           steps=[
               ("GET /storages/names/crews?parentId={unitId}", "HTTP 200"),
               ("Перевірити наявність crew.id", "crew у списку"),
           ]),
        mk("TC-STR-CREW-013", FEAT_CREW_REG, "AC-09", "getCrewNames — рекурсивний пошук у дочірніх UNIT",
           "Екіпаж з вкладеного UNIT-AB видимий при пошуку від батьківського UNIT-A.",
           preconditions=PRE_ADMIN, tags=api,
           test_data="Ієрархія UNIT-A → UNIT-AB → CREW.",
           steps=[
               ("Підготувати hierarchy scenario", "unitA, unitAB, crew"),
               ("GET crew names для UNIT-A", "crew.id присутній"),
           ]),
        mk("TC-STR-CREW-014", FEAT_CREW_REG, "AC-10", "getCrewUnits — дерево з коренем UNIT-A",
           "Combobox підрозділів показує батьківський UNIT, не лише leaf nodes.",
           preconditions=PRE_ADMIN, tags=api,
           steps=[("GET crew-units для hierarchy", "root містить unitA.id; child unitAB")]),
    ]


def crew_relocation_cases() -> list[Case]:
    amt = "ISSUE_AMOUNT=15; stock sender=100."
    return [
        mk("TC-CREW-REL-001", FEAT_CREW_ISSUE, "AC-01", "Видача UNIT→CREW — CREATED→FINISHED відправником",
           "Send на CREW спочатку CREATED («В дорозі»); зарахування на екіпаж після підтвердження відправником.",
           priority="CRITICAL", preconditions=PRE_CREW, tags="crew,relocations,api",
           test_data=amt, role_name=ROLE_OWNER,
           steps=[
               ("Зафіксувати stock sender і crew", "before"),
               ("POST send UNIT → CREW", "state=CREATED; sender −N; crew без +N"),
               ("PUT resolve FINISHED відправником", "state=FINISHED; crew +N"),
           ]),
        mk("TC-CREW-REL-002", FEAT_CREW_ISSUE, "AC-02", "Журнал: «В дорозі» після send, «Видано» після finish",
           "Після send рядок у CREATED; після FINISHED — у sent history.",
           preconditions=PRE_CREW, tags="crew,relocations,api", role_name=ROLE_OWNER,
           steps=[
               ("POST send з маркером", "CREATED у in-transit"),
               ("PUT resolve FINISHED", "FINISHED"),
               ("GET sent history", "relocationId у content"),
           ]),
        mk("TC-CREW-REL-003", FEAT_CREW_ISSUE, "AC-03", "Негатив: недостатній stock при видачі на CREW",
           "Не можна видати більше, ніж є на складі відправника.",
           preconditions=PRE_CREW, tags="crew,relocations,validation", role_name=ROLE_OWNER,
           steps=[
               ("POST send amount > stock", "HTTP 400"),
               ("Перевірити stock", "без змін"),
           ]),
        mk("TC-CREW-REL-004", FEAT_CREW_ISSUE, "AC-04", "Multi-resource send (2+ позиції)",
           "Одна видача може містити кілька ресурсів; credit на crew після FINISHED.",
           preconditions=PRE_CREW, tags="crew,relocations,api", role_name=ROLE_OWNER,
           steps=[
               ("Підготувати 2 ресурси з stock", "resource A, B"),
               ("POST send з 2 рядками", "CREATED"),
               ("PUT resolve FINISHED відправником", "crew +N для кожного"),
           ]),
        mk("TC-CREW-REL-005", FEAT_CREW_ISSUE, "AC-05", "OWNER_2 не може send на crew поза CREWS region",
           "RBAC області CREWS: чужий підрозділ не бачить чужі екіпажі.",
           preconditions=PRE_OWNER2, tags="crew,relocations,rbac", role_name=ROLE_OWNER,
           steps=[
               ("Спроба send на crew з OWNER_1 сценарію під OWNER_2", "HTTP 403 або 404"),
               ("Перевірити stock", "без змін"),
           ]),
        mk("TC-CREW-REL-006", FEAT_CREW_ISSUE, "AC-06", "Видача з PRODUCTION sender",
           "Виробнича локація: CREATED → FINISHED відправником.",
           preconditions=PRE_CREW, tags="crew,relocations,api", role_name=ROLE_OWNER,
           test_data="Ephemeral PRODUCTION child; stock; crew recipient.",
           steps=[
               ("POST send PRODUCTION → CREW", "CREATED"),
               ("PUT resolve FINISHED відправником", "crew +N"),
           ]),
        mk("TC-CREW-REL-007", FEAT_CREW_ISSUE, "AC-07", "Журнал отримувача (crew storage)",
           "Після FINISHED видача видима в received history crew storage.",
           preconditions=PRE_CREW, tags="crew,relocations,api", role_name=ROLE_OWNER,
           steps=[
               ("Send + finish UNIT→CREW", "FINISHED"),
               ("GET received history для crew.id", "relocationId у content"),
           ]),
        mk("TC-CREW-REL-008", FEAT_CREW_ISSUE, "AC-08", "Multi-item — кожна позиція ≤ stock",
           "Валідація по кожному рядку: сума не перевищує залишок; lifecycle CREATED→FINISHED.",
           preconditions=PRE_CREW, tags="crew,relocations,api", role_name=ROLE_OWNER,
           steps=[("POST send 2 resources у межах stock + finish", "CREATED→FINISHED")]),
        mk("TC-CREW-REL-009", FEAT_CREW_ISSUE, "AC-09", "UNIT→CREW прихований для ACCOUNTANT",
           "Бухгалтер не бачить внутрішні видачі на екіпажі в GET /relocations.",
           preconditions=PRE_CREW, tags="crew,relocations,accountant",
           role_name=ROLE_ACC,
           steps=[
               ("OWNER_1: send+finish UNIT→CREW", "relocationId"),
               ("GET /relocations як ACCOUNTANT", "relocationId відсутній"),
           ]),
        mk("TC-CREW-REL-010", FEAT_CREW_ISSUE, "AC-23", "Отримувач CREW не може FINISHED",
           "Підтвердження доставки на CREW виконує лише відправник.",
           priority="CRITICAL", preconditions=PRE_CREW, tags="crew,relocations,api",
           role_name=ROLE_OWNER,
           steps=[
               ("POST send → CREATED", "relocationId"),
               ("resolve FINISHED зі storageId=crew", "HTTP 4xx"),
           ]),
        mk("TC-CREW-REL-011", FEAT_CREW_ISSUE, "AC-24", "Відправник скасовує CREATED (RETURNED)",
           "Скасування лише відправником відновлює stock.",
           priority="CRITICAL", preconditions=PRE_CREW, tags="crew,relocations,api",
           role_name=ROLE_OWNER,
           steps=[
               ("POST send → CREATED", "sender −N"),
               ("resolve RETURNED відправником", "stock відновлено"),
           ]),
        mk("TC-CREW-HIST-001", FEAT_CREW_ISSUE, "AC-10", "Картка «Видано» — totalRemovedResources +N",
           "Після send (CREATED) агрегат removed resources на sender збільшується на amount.",
           preconditions=PRE_CREW, tags="crew,history,api", role_name=ROLE_OWNER,
           steps=[
               ("Зафіксувати totalRemovedResources для resource", "before"),
               ("POST send UNIT→CREW", "CREATED"),
               ("Перевірити totalRemovedResources", "delta ≈ ISSUE_AMOUNT"),
           ]),
        mk("TC-FLY-REL-001", FEAT_CREW_ISSUE, "AC-18", "Send на FLY_POINT — CREATED→FINISHED",
           "Видача на точку вильоту: підтверджує відправник; баланс на FLY_POINT після FINISHED.",
           priority="CRITICAL", preconditions=PRE_CREW, tags="crew,fly-point,relocations,api",
           role_name=ROLE_OWNER,
           steps=[
               ("POST send → FLY_POINT", "CREATED"),
               ("PUT resolve FINISHED відправником", "FLY_POINT +N"),
           ]),
        mk("TC-FLY-REL-002", FEAT_CREW_ISSUE, "AC-19", "Attached CREW — auto-forward на FLY_POINT",
           "Якщо parent екіпажу = FLY_POINT, після FINISHED залишок опиняється на точці.",
           priority="CRITICAL", preconditions=PRE_CREW, tags="crew,fly-point,relocations,api",
           role_name=ROLE_OWNER,
           steps=[
               ("Send → attached CREW → FINISHED", "FINISHED"),
               ("Перевірити stock", "crew без +N; FLY_POINT +N"),
           ]),
        mk("TC-FLY-REL-003", FEAT_CREW_ISSUE, "AC-20", "Пряме CREW→FLY_POINT — AUTO_FINISHED",
           "Переміщення з екіпажу на точку вильоту миттєве.",
           priority="CRITICAL", preconditions=PRE_CREW, tags="crew,fly-point,relocations,api",
           role_name=ROLE_OWNER,
           steps=[
               ("Seed stock на CREW (send+finish)", "crew +N"),
               ("POST CREW → FLY_POINT", "AUTO_FINISHED; crew −N; FLY_POINT +N"),
           ]),
        mk("TC-CREW-INC-001", FEAT_CREW_ISSUE, "AC-21", "Надзвичайна подія на send→CREW",
           "Incident → LOST; ресурс списаний з відправника, на CREW не зараховується.",
           priority="CRITICAL", preconditions=PRE_CREW, tags="crew,incident,api",
           role_name=ROLE_OWNER,
           steps=[
               ("POST send → CREW", "CREATED"),
               ("POST incident", "LOST; crew без credit; WRITE_OFF на sender"),
           ]),
        mk("TC-CREW-INC-002", FEAT_CREW_ISSUE, "AC-22", "Надзвичайна подія на send→FLY_POINT",
           "Incident → LOST; FLY_POINT без credit.",
           priority="CRITICAL", preconditions=PRE_CREW, tags="crew,fly-point,incident,api",
           role_name=ROLE_OWNER,
           steps=[
               ("POST send → FLY_POINT + incident", "LOST; FLY_POINT без +N"),
           ]),
    ]


def crew_inventory_cases() -> list[Case]:
    return [
        mk("TC-CREW-INV-001", FEAT_CREW_INV, "AC-01", "Звіт STOCK — залишки на екіпажі після видачі",
           "GET /storages/inventory/crews?requestType=STOCK показує amount після видачі UNIT→CREW.",
           priority="CRITICAL", preconditions=PRE_CREW, tags="crew,inventory,api",
           test_data="Після send amount=20; crew+resource у звіті.",
           role_name=ROLE_OWNER,
           steps=[
               ("OWNER_1: send+finish UNIT→CREW amount=20", "FINISHED; crew +20"),
               ("GET /inventory/crews requestType=STOCK", "HTTP 200"),
               ("Знайти рядок crew+resource", "amount≈20"),
           ]),
        mk("TC-CREW-INV-002", FEAT_CREW_INV, "AC-02", "Звіт INCOME — сума видач за період",
           "Агрегат надходжень на екіпаж за fromDate/toDate включає сьогоднішню видачу.",
           preconditions=PRE_CREW, tags="crew,inventory,api", role_name=ROLE_OWNER,
           test_data="fromDate/toDate = сьогодні ±1 день.",
           steps=[
               ("Після send: GET /inventory/crews requestType=INCOME", "HTTP 200"),
               ("Перевірити income для crew+resource", "income ≥ ISSUE_AMOUNT"),
           ]),
        mk("TC-CREW-INV-006", FEAT_CREW_INV, "AC-03", "STOCK-звіт = direct inventory (Crew-Manager)",
           "Звіт /inventory/crews узгоджений з GET /storages/{crewId}/inventory для Crew-Manager.",
           preconditions=PRE_CM, tags="crew,inventory,api", role_name=ROLE_OWNER,
           steps=[
               ("GET STOCK report", "amount_report"),
               ("GET /storages/{crewId}/inventory як CREW_MANAGER", "amount_direct"),
               ("Порівняти", "різниця < 0.01"),
           ]),
        mk("TC-CREW-INV-007", FEAT_CREW_INV, "AC-04", "OWNER_1 — direct crew inventory заборонено (403)",
           "Business Unit Owner без ролі Crew-Manager не читає inventory-list екіпажу напряму.",
           preconditions=PRE_CREW, tags="crew,inventory,rbac", role_name=ROLE_OWNER,
           steps=[("GET /storages/{crewId}/inventory як OWNER_1", "HTTP 403")]),
        mk("TC-CREW-INV-007b", FEAT_CREW_INV, "AC-05", "Crew-Manager читає direct crew inventory",
           "Користувач argument (Crew-Manager) має inventory-list::{crew}::read.",
           priority="CRITICAL", preconditions=PRE_CM, tags="crew,inventory,api", role_name=ROLE_OWNER,
           steps=[
               ("GET /storages/{crewId}/inventory", "HTTP 200"),
               ("Перевірити stock resource", "≈ ISSUE_AMOUNT"),
           ]),
        mk("TC-CREW-INV-008", FEAT_CREW_INV, "AC-06", "OWNER_2 поза CREWS region — 403/404",
           "Чужий підрозділ не бачить inventory екіпажу з чужої області.",
           preconditions=PRE_OWNER2, tags="crew,inventory,rbac", role_name=ROLE_OWNER,
           steps=[("GET /storages/{crewId}/inventory як OWNER_2", "HTTP 403 або 404")]),
        mk("TC-CREW-INV-009", FEAT_CREW_INV, "AC-07", "Відкриття inventory session: ADMIN ✓, OWNER_2 ✗",
           "Лише адмін або уповноважений Crew-Manager керує сесією інвентаризації екіпажу.",
           preconditions=PRE_CREW, tags="crew,inventory,rbac",
           steps=[
               ("PUT /inventory/status open як ADMIN", "HTTP 200; open=true"),
               ("Те саме як OWNER_2", "HTTP 403"),
           ]),
        mk("TC-CREW-INV-010", FEAT_CREW_INV, "AC-08", "PUT inventory змінює amount на crew storage",
           "Проведена інвентаризація оновлює залишок на балансі екіпажу.",
           preconditions=PRE_CREW, tags="crew,inventory,api",
           steps=[
               ("ADMIN: open session на crew", "session open"),
               ("PUT /inventory target amount", "HTTP 200"),
               ("conduct; перевірити stock", "amount = target"),
           ]),
        mk("TC-CREW-FIGHT-001", FEAT_CREW_INV, "AC-09", "Fight sync — write-off у журналі (dev integration)",
           "Після синхронізації з Fight з'являється запис списання для екіпажу. Потребує Fight на dev.",
           preconditions="Fight integration enabled на dev.", tags="crew,fight,manual-only", layer=None,
           steps=[("Після Fight sync: GET /write-off page для crew", "Запис присутній")]),
        mk("TC-CREW-FIGHT-002", FEAT_CREW_INV, "AC-09", "Fight — complete reconciliation зменшує stock",
           "Завершення звірки списує ресурси з балансу екіпажу. Потребує Fight seed.",
           preconditions="Fight integration + seed на dev.", tags="crew,fight,manual-only", layer=None,
           steps=[("Complete reconciliation", "crew stock зменшився")]),
        mk("TC-CREW-WO-PROBE", FEAT_CREW_INV, "AC-10", "GET /write-off/short-stats — contract probe",
           "Перевірка доступності API списань для OWNER_1 (200 або 403 за RBAC).",
           severity="MINOR", preconditions=PRE_OWNER1, tags="crew,write-off,api",
           role_name=ROLE_OWNER,
           steps=[("GET /write-off/short-stats як OWNER_1", "HTTP 200 або 403")]),
    ]


def crew_ui_cases() -> list[Case]:
    pre = f"{PRE_OWNER1} Область CREWS налаштована; на member stock ресурсу ≥ 100 од."
    return [
        mk("TC-UI-CREW-001", FEAT_CREW_ISSUE, "AC-11", "UI — кнопка «Видати на екіпаж» видима",
           "На /relocations для member storage з CREWS region відображається кнопка видачі на екіпаж.",
           priority="CRITICAL", preconditions=pre, tags="crew,ui", role_name=ROLE_OWNER, layer="UI",
           steps=[
               ("Відкрити /relocations під OWNER_1 на member storage", "Сторінка завантажена"),
               ("Знайти кнопку «Видати на екіпаж»", "Кнопка видима"),
           ]),
        mk("TC-UI-CREW-002", FEAT_CREW_ISSUE, "AC-12", "UI — happy path: «В дорозі» → finish → «Видано»",
           "Повний сценарій: підрозділ → екіпаж → submit → «В дорозі»; finish відправником → «Видано».",
           priority="CRITICAL", preconditions=pre, tags="crew,ui", role_name=ROLE_OWNER, layer="UI",
           test_data="Кількість=10; видав: «UI Тест Видав», звання «Сержант».",
           steps=[
               ("Натиснути «Видати на екіпаж»", "Форма відкрита"),
               ("Обрати UNIT → CREW → ресурс; qty=10; заповнити видав", "Кнопка «Підтвердити» активна"),
               ("Підтвердити", "Рядок у «В дорозі»"),
               ("Finish відправником (API/UI)", "FINISHED"),
               ("Вкладка «Видано» — знайти ресурс/екіпаж", "Рядок присутній"),
           ]),
        mk("TC-UI-CREW-003", FEAT_CREW_ISSUE, "AC-13", "UI — «Всі локації» приховує кнопку",
           "У режимі «Всі локації» write-операції заблоковані — кнопка видачі на екіпаж не показується.",
           preconditions=pre, tags="crew,ui,guard", role_name=ROLE_OWNER, layer="UI",
           steps=[
               ("Встановити selectedStorageId=all (localStorage)", "Режим «Всі локації»"),
               ("Відкрити /relocations", "Сторінка завантажена"),
               ("Перевірити кнопку «Видати на екіпаж»", "Кнопка прихована"),
           ]),
        mk("TC-UI-CREW-004", FEAT_CREW_INV, "AC-11", "UI — залишки екіпажу на /unit-management",
           "Після видачі таблиця залишків екіпажу показує amount (режим «Екіпажі»).",
           preconditions=pre, tags="crew,ui,stock", role_name=ROLE_OWNER, layer="UI",
           steps=[
               ("Виконати видачу UNIT→CREW send+finish (API або UI)", "stock на crew"),
               ("Відкрити /unit-management → режим екіпажів", "Таблиця завантажена"),
               ("Обрати екіпаж; знайти ресурс", "amount ≈ видача"),
           ]),
        mk("TC-UI-CREW-005", FEAT_CREW_ISSUE, "AC-14", "UI — каскад UNIT → екіпаж у combobox",
           "Після вибору батьківського UNIT список екіпажів не порожній (ієрархія).",
           priority="CRITICAL", preconditions=pre, tags="crew,ui", role_name=ROLE_OWNER, layer="UI",
           test_data="Hierarchy: UNIT-A → UNIT-AB → CREW.",
           steps=[
               ("Відкрити форму видачі на екіпаж", "Форма"),
               ("Обрати батьківський UNIT-A", "UNIT обрано"),
               ("Відкрити combobox екіпажів", "Список не порожній; є crew з дочірнього UNIT"),
           ]),
        mk("TC-UI-CREW-006", FEAT_CREW_ISSUE, "AC-15", "UI — OWNER_2 без membership не бачить екіпажі OWNER_1",
           "Кнопка/форма недоступні для підрозділу поза областю CREWS.",
           preconditions=PRE_OWNER2, tags="crew,ui,rbac", role_name=ROLE_OWNER, layer="UI",
           steps=[
               ("Відкрити /relocations під OWNER_2", "Сторінка"),
               ("Перевірити кнопку «Видати на екіпаж»", "Прихована або форма без екіпажів OWNER_1"),
           ]),
        mk("TC-UI-CREW-007", FEAT_CREW_ISSUE, "AC-16", "UI — валідація кількості > stock",
           "Кнопка підтвердження неактивна при перевищенні залишку; активна при валідній qty.",
           preconditions=pre, tags="crew,ui,validation", role_name=ROLE_OWNER, layer="UI",
           steps=[
               ("Ввести qty > stock (напр. 9999)", "«Підтвердити» disabled"),
               ("Ввести qty=10", "«Підтвердити» enabled"),
           ]),
        mk("TC-UI-CREW-009", FEAT_CREW_ISSUE, "AC-17", "UI — картка «Видано» в історії операцій",
           "Після видачі на екіпаж зростає показник removed на sender у картці ресурсу.",
           preconditions=pre, tags="crew,ui,history", role_name=ROLE_OWNER, layer="UI",
           steps=[
               ("Після send: відкрити історію операцій / картку ресурсу", "Сторінка"),
               ("Знайти summary «Видано»", "Значення збільшилось на amount видачі"),
           ]),
        mk("TC-UI-CREW-010", FEAT_CREW_INV, "AC-12", "UI — режим екіпажів без обраного crew",
           "До вибору екіпажу таблиця залишків порожня / placeholder.",
           preconditions=pre, tags="crew,ui,stock", role_name=ROLE_OWNER, layer="UI",
           steps=[
               ("Відкрити /unit-management → екіпажі", "Режим активний"),
               ("Не обирати екіпаж", "Таблиця порожня або prompt «Оберіть екіпаж»"),
           ]),
        mk("TC-UI-CREW-011", FEAT_CREW_INV, "AC-13", "UI — toggle інвентаризації після вибору crew",
           "Кнопки інвентаризації стають доступними лише після вибору конкретного екіпажу.",
           preconditions=pre, tags="crew,ui,inventory", role_name=ROLE_OWNER, layer="UI",
           steps=[
               ("Режим екіпажів без вибору", "Toggle інвентаризації disabled"),
               ("Обрати екіпаж з CREWS region", "Toggle enabled"),
           ]),
    ]


def relocation_journal_cases() -> list[Case]:
    pre = f"{PRE_OWNER1} У журналі є записи з різними відправниками/отримувачами; stock ресурсу seed."
    return [
        mk("TC-REL-007", FEAT_WMS_JOURNAL, "AC-11", "Журнал «Отримано» — фільтр категорії + сорт «До» ASC",
           "Комбінація фільтра categoryId і сортування за іменем отримувача (recipient.name ASC).",
           priority="CRITICAL", preconditions=pre, tags="relocations,filters,api", role_name=ROLE_OWNER,
           steps=[
               ("GET /relocations received history + categoryId + sort=recipient.name ASC", "HTTP 200"),
               ("Перевірити порядок recipient.name", "ASC; опорний запис у результаті"),
           ]),
        mk("TC-REL-008", FEAT_WMS_JOURNAL, "AC-11", "Журнал «Отримано» — категорія + сорт «Від» ASC",
           "Сортування за sender.name при активному фільтрі категорії.",
           priority="CRITICAL", preconditions=pre, tags="relocations,filters,api", role_name=ROLE_OWNER,
           steps=[("GET received + categoryId + sort=sender.name ASC", "Порядок sender.name ASC")]),
        mk("TC-REL-009", FEAT_WMS_JOURNAL, "AC-11", "Журнал «Отримано» — продукт + сорт «До» DESC",
           "Фільтр productId + сортування recipient.name DESC.",
           priority="CRITICAL", preconditions=pre, tags="relocations,filters,api", role_name=ROLE_OWNER,
           steps=[("GET received + productId + sort=recipient.name DESC", "DESC order")]),
        mk("TC-REL-047", FEAT_WMS_JOURNAL, "AC-11", "Журнал «Видано» — категорія + сорт «Від» ASC",
           "Вкладка sent history: category filter + sender.name ASC.",
           priority="CRITICAL", preconditions=pre, tags="relocations,filters,api", role_name=ROLE_OWNER,
           steps=[("GET sent + categoryId + sort=sender.name ASC", "Порядок коректний")]),
        mk("TC-REL-048", FEAT_WMS_JOURNAL, "AC-11", "Журнал «Видано» — продукт + сорт «До» DESC",
           "Sent history: product filter + recipient.name DESC.",
           priority="CRITICAL", preconditions=pre, tags="relocations,filters,api", role_name=ROLE_OWNER,
           steps=[("GET sent + productId + sort=recipient.name DESC", "DESC order")]),
        mk("TC-UI-REL-009", FEAT_WMS_JOURNAL, "AC-11", "UI — фільтр продукту + сортування «До» на /relocations",
           "Ручна перевірка: фільтр продукту на UI відповідає API; сортування за отримувачем ASC.",
           preconditions=pre, tags="relocations,ui,filters", role_name=ROLE_OWNER, layer="UI",
           steps=[
               ("Відкрити /relocations → «Видано»", "Журнал"),
               ("Застосувати фільтр «Продукт»", "Таблиця оновилась"),
               ("Сортувати за колонкою «До» ASC", "Порядок відповідає API"),
           ]),
    ]


def relocation_invoice_ui_cases() -> list[Case]:
    pre_api = (
        f"{PRE_ADMIN} OWNER_2 accessMode=REGIONS. FULL_ACCESS region з ephemeral STORAGE child. "
        "Stock на sender; generateInvoice=true."
    )
    return [
        mk("TC-UI-REL-011", FEAT_WMS_INVOICE, "AC-01", "API+UI — накладна від STORAGE sender",
           "Відправник type=STORAGE: generateInvoice=true → № у журналі; download PDF/DOCX >100 bytes.",
           priority="CRITICAL", preconditions=pre_api, tags="relocations,invoice,api",
           test_data="ADMIN створює send; OWNER_2 перевіряє journal + GET /invoice.",
           steps=[
               ("ADMIN: POST /send?generateInvoice=true (STORAGE sender)", "canGenerateInvoice=true"),
               ("GET journal «В дорозі»/«Видано»", "invoiceNumber присутній"),
               ("GET /relocations/{id}/invoice", "Файл >100 bytes"),
           ]),
        mk("TC-UI-REL-012", FEAT_WMS_INVOICE, "AC-02", "UI — отримувач завантажує накладну з «Отримано»",
           "Отримувач (OWNER_2) бачить № накладної і завантажує файл з вкладки «Отримано».",
           priority="CRITICAL",
           preconditions=f"{pre_api} API: send FINISHED; UI під OWNER_2 workspace=recipient.",
           tags="relocations,invoice,ui", role_name=ROLE_OWNER, layer="UI",
           steps=[
               ("API setup: send + resolve FINISHED", "invoiceNumber у received history"),
               ("UI OWNER_2: /relocations → «Отримано»", "Рядок з № накладної"),
               ("Клік завантажити накладну", "Download PDF/DOCX успішний"),
           ]),
        mk("TC-UI-REL-013", FEAT_WMS_INVOICE, "AC-01", "API — накладна від PRODUCTION sender",
           "Те саме що TC-UI-REL-011 для sender type=PRODUCTION (relation=INTERNAL).",
           preconditions=pre_api, tags="relocations,invoice,api",
           test_data="Ephemeral PRODUCTION child; generateInvoice=true.",
           steps=[
               ("ADMIN: send PRODUCTION→recipient generateInvoice=true", "invoiceNumber"),
               ("GET /invoice на «В дорозі» та «Видано»", "Download OK"),
           ]),
        mk("TC-UI-REL-014", FEAT_WMS_INVOICE, "AC-03", "UI — UNIT (ПМ БАР) завантажує накладну",
           "Відправник-підрозділ завантажує накладну з журналу «В дорозі» та «Видано» на UI.",
           priority="CRITICAL",
           preconditions=f"{PRE_OWNER2} FULL_ACCESS region; видача з ПМ БАР → recipient у locations.",
           tags="relocations,invoice,ui", role_name=ROLE_OWNER, layer="UI",
           steps=[
               ("OWNER_2: POST send generateInvoice=true (sender=owner2 unit)", "invoiceNumber"),
               ("UI: журнал «В дорозі» — клік № накладної", "Playwright download >100 bytes"),
               ("Повторити на вкладці «Видано»", "Download OK"),
           ]),
    ]


def visibility_region_case() -> list[Case]:
    return [
        mk("TC-STR-REG-032", FEAT_REGION_LOC, "AC-19", "Union видимості — member у двох областях",
           "Підрозділ-member у двох FULL_ACCESS областях бачить об'єднання loc1 ∪ loc2 у GET /names.",
           preconditions=PRE_ADMIN, tags="visibility-regions,api",
           test_data="region1→loc1, region2→loc2; OWNER_2 member в обох.",
           steps=[
               ("Створити 2 області з різними locations", "region1, region2"),
               ("Додати OWNER_2 member до обох", "setup OK"),
               ("GET /storages/names як OWNER_2", "Містить loc1.id і loc2.id"),
           ]),
    ]


def resource_viewer_cases() -> list[Case]:
    return [
        mk("TC-RVW-API-001", FEAT_RES_RBAC, "AC-01", "API — my-units без UNIT для ResourceViewer",
           "Роль Resource Viewer бачить у селекторі лише STORAGE/PRODUCTION, не підрозділи UNIT.",
           priority="CRITICAL",            preconditions="@ResourceViewer (wolf) залогінений.", tags="resource-viewer,api", role_name=ROLE_ADMIN,
           steps=[
               ("GET /storages/names/my-units як RESOURCE_VIEWER", "HTTP 200"),
               ("Перевірити: жоден type=UNIT", "Список без UNIT"),
           ]),
        mk("TC-RVW-API-002", FEAT_RVW, "AC-01", "API — фільтр sender STORAGE/PRODUCTION only",
           "Журнал переміщень Resource Viewer: senderIds лише operational storages.",
           preconditions="@ResourceViewer; fixture relocations з STORAGE/PRODUCTION senders.",
           tags="resource-viewer,api", role_name=ROLE_ADMIN,
           steps=[
               ("GET /resources-viewer/relocations з sender filter", "HTTP 200"),
               ("Перевірити sender.type", "Лише STORAGE або PRODUCTION"),
           ]),
        mk("TC-RVW-API-003", FEAT_RVW, "AC-02", "API — crew/unit send прихований у viewer journal",
           "Видачі UNIT→CREW та UNIT→UNIT не потрапляють у viewer relocation journal.",
           preconditions="@ResourceViewer; fixture UNIT→CREW + STORAGE→UNIT.",
           tags="resource-viewer,api", role_name=ROLE_ADMIN,
           steps=[
               ("Fixture: UNIT→CREW і STORAGE→UNIT", "2 relocationId"),
               ("GET viewer relocations", "Є STORAGE→UNIT; немає UNIT→CREW"),
           ]),
        mk("TC-RVW-001", FEAT_RVW, "AC-03", "API — relocations/sum сортування resourceName ASC",
           "Порядок у зведенні: цифри → латиниця → укр. л/є/і/ї (natural order).",
           preconditions="@ResourceViewer (wolf) залогінений.", tags="resource-viewer,api", role_name=ROLE_ADMIN,
           test_data="8 ресурсів: 111_rvw_, aaa_rvw_, …, їжа_rvw_.",
           steps=[
               ("Створити 8 тестових ресурсів з префіксами", "ids"),
               ("GET /resources-viewer/relocations/sum?resourceIds=…&receiverIds=", "HTTP 200"),
               ("Перевірити порядок resourceName", "111 → aaa → … → їжа"),
           ]),
        mk("TC-UI-RES-AC-001", FEAT_RES_RBAC, "AC-02", "UI — autocomplete ресурсів з фільтром категорії",
           "Resource Viewer на сторінці перегляду ресурсів фільтрує autocomplete за categoryIds.",
           preconditions="@ResourceViewer (wolf) залогінений.", tags="resource-viewer,ui",
           role_name=ROLE_ADMIN, layer="UI",
           steps=[
               ("Відкрити сторінку resource viewer", "Сторінка завантажена"),
               ("Обрати категорію A у фільтрі", "Autocomplete оновився — лише ресурси категорії A"),
               ("Скинути фільтр", "Повний список за search"),
           ]),
    ]


def tech_map_plan_guard_cases() -> list[Case]:
    pre = f"{PRE_ADMIN} OWNER_1 storage: mode=EDIT_ALLOWED; ізольована PRODUCTION техкарта."
    return [
        mk("TC-MFG-026", FEAT_MFG_TM_ARCH, "AC-01", "Заборона деактивації єдиної техкарти в активному плані",
           "Не можна заархівувати єдину техкарту продукту, якщо продукт у плані поточного/майбутнього місяця.",
           priority="CRITICAL", preconditions=pre, tags="tech-maps,plans,validation",
           steps=[
               ("Створити tech map + location plan на наступний місяць з цим продуктом", "planId"),
               ("DELETE /technological-maps/{id} (deactivate)", "HTTP 400 used-in-plan"),
               ("Перевірити: техкарта active", "count active без змін"),
           ]),
        mk("TC-MFG-027", FEAT_MFG_TM_ARCH, "AC-01", "Заборона деактивації — продукт лише у майбутньому плані",
           "План на +3 місяці також блокує деактивацію єдиної техкарти.",
           priority="CRITICAL", preconditions=pre, tags="tech-maps,plans,validation",
           steps=[
               ("Plan на +3 місяці з продуктом tech map", "planId"),
               ("DELETE deactivate sole tech map", "HTTP 400"),
           ]),
        mk("TC-MFG-028", FEAT_MFG_TM_ARCH, "AC-01", "Деактивація дозволена при альтернативній активній техкарті",
           "Якщо для продукту є інша active tech map — першу можна заархівувати навіть при плані.",
           priority="CRITICAL", preconditions=pre, tags="tech-maps,plans,api",
           steps=[
               ("2 active tech maps на один продукт; plan з продуктом", "setup"),
               ("DELETE deactivate одну з карт", "HTTP 200"),
               ("Друга карта лишається active", "count active ≥ 1"),
           ]),
        mk("TC-MFG-029", FEAT_MFG_TM_ARCH, "AC-01", "Деактивація дозволена без плану на продукт",
           "Єдина техкарта деактивується, якщо продукт не в жодному актуальному плані.",
           preconditions=pre, tags="tech-maps,api",
           steps=[
               ("Tech map без plan reference", "id"),
               ("DELETE deactivate", "HTTP 200; active=false"),
           ]),
        mk("TC-MFG-030", FEAT_MFG_TM_ARCH, "AC-01", "Owner і Admin — однакова заборона при plan guard",
           "RBAC: і OWNER_1 і ADMIN отримують 400 при спробі деактивувати sole map у плані.",
           preconditions=pre, tags="tech-maps,plans,rbac",
           steps=[
               ("Plan + sole tech map", "setup"),
               ("DELETE як OWNER_1", "HTTP 400"),
               ("DELETE як ADMIN", "HTTP 400"),
           ]),
    ]


def global_plan_and_plans_cases() -> list[Case]:
    return [
        mk("TC-GP-045", FEAT_GP, "AC-GP-09", "Декомпозиція після деактивації техкарти зі snapshot",
           "Відомий дефект: POST /decompose зі збереженим assignment на deactivated map → 400. "
           "Очікувана поведінка після фіксу: 200 з alternate map. Тест відтворює UI wizard Tab 2.",
           priority="CRITICAL",
           preconditions=f"{PRE_ADMIN} Global plan з decomposition snapshot; tech map A deactivated, B active.",
           tags="global-plans,decompose,regression",
           steps=[
               ("Створити global plan + generate decomposition з map A", "snapshot збережено"),
               ("Деактивувати tech map A", "active=false"),
               ("POST /decompose block 0 з assignments з snapshot (як UI start)", "Зараз: HTTP 400; після фіксу: 200"),
           ]),
        mk("TC-GP-046", FEAT_GP, "AC-GP-09", "DELETE техкарти заблоковано глобальним планом (майбутній місяць)",
           "Після generate snapshot містить technologicalMapId; DELETE /technological-maps/{id} → HTTP 400.",
           priority="CRITICAL",
           preconditions=f"{PRE_ADMIN} Global plan chain M1/M2/M3; EDIT_ALLOWED на L1.",
           tags="global-plans,tech-maps,guard",
           steps=[
               ("POST global-plans + decompose + generate з M1", "decomposition snapshot збережено"),
               ("DELETE M1 @ L1", "HTTP 400; повідомлення про глобальний план"),
           ]),
        mk("TC-GP-047", FEAT_GP, "AC-GP-09", "DELETE техкарти заблоковано глобальним планом (поточний місяць)",
           "Guard спрацьовує для gp.to_date >= початок поточного місяця.",
           priority="CRITICAL",
           preconditions=f"{PRE_ADMIN} Вільний поточний календарний місяць.",
           tags="global-plans,tech-maps,guard",
           steps=[
               ("Generate global plan на поточний місяць з M1", "snapshot збережено"),
               ("DELETE M1", "HTTP 400"),
           ]),
        mk("TC-GP-048", FEAT_GP, "AC-GP-09", "Global plan guard не знімається при alternate tech map",
           "На відміну від per-location (TC-MFG-028): snapshot посилається на M1 — DELETE M1 → 400 навіть з другою активною карткою.",
           priority="CRITICAL",
           preconditions=f"{PRE_ADMIN} Дві активні техкарти на ресурс A.",
           tags="global-plans,tech-maps,guard",
           steps=[
               ("Alternate tech map + generate з M1", "snapshot з M1"),
               ("DELETE M1", "HTTP 400"),
           ]),
        mk("TC-GP-049", FEAT_GP, "AC-GP-09", "DELETE дозволено без generate snapshot",
           "Лише POST /global-plans без generate — decomposition=null, guard не блокує.",
           preconditions=f"{PRE_ADMIN} Ізольована техкарта.",
           tags="global-plans,tech-maps,guard",
           steps=[
               ("POST global-plans без generate", "decomposition=null"),
               ("DELETE isolated tech map", "HTTP 200"),
           ]),
        mk("TC-PLAN-001", FEAT_PLAN, "AC-03", "Admin «Всі локації» — сортування планів from DESC, name ASC",
           "На /plans без storageId список сортується: дата початку (from) спадання, потім назва локації А-Я.",
           priority="CRITICAL", preconditions=PRE_ADMIN, tags="plans,api,sort",
           steps=[
               ("Створити плани на L1/L2 з різними month/year", "3 plan id"),
               ("GET /plans без storageId (admin all locations)", "HTTP 200"),
               ("Перевірити порядок: newer from першими; при рівній даті — storage.name ASC", "Sort OK"),
           ]),
    ]


def equipment_cases() -> list[Case]:
    return [
        mk("TC-EQ-SEL-001", FEAT_EQU, "AC-01", "API — «Звідки» лише EXTERNAL (+ SUPPLIER)",
           "Dropdown «Звідки» на /equipment фільтрує relation=EXTERNAL, не type.",
           priority="CRITICAL", preconditions=PRE_ADMIN, tags="equipment,api,relation",
           steps=[
               ("Створити INTERNAL і EXTERNAL child STORAGE", "2 id"),
               ("GET /names?relation=EXTERNAL", "EXTERNAL є; INTERNAL немає; є SUPPLIER"),
           ]),
        mk("TC-EQ-SEL-002", FEAT_EQU, "AC-01", "Створення обладнання з EXTERNAL supplier",
           "Позитивний шлях: sender=SUPPLIER (EXTERNAL), recipient=owner storage.",
           priority="CRITICAL", preconditions=PRE_ADMIN, tags="equipment,api",
           steps=[
               ("POST /equipment senderStorageId=supplier, storage=owner1", "HTTP 200/201"),
               ("GET equipment", "storage.id = owner1"),
           ]),
        mk("TC-EQ-SEL-003", FEAT_EQU, "AC-01", "Негатив: INTERNAL STORAGE як sender",
           "Валідація забороняє INTERNAL STORAGE як джерело обладнання.",
           preconditions=PRE_ADMIN, tags="equipment,validation,api",
           steps=[("POST equipment senderStorageId=INTERNAL child", "HTTP 400 field=senderStorageId")]),
        mk("TC-UI-EQ-001", FEAT_EQU, "AC-01", "UI — фільтр «Відповідальний» на /equipment",
           "Журнал обладнання фільтрується за співробітником у dropdown assignee.",
           preconditions=f"{PRE_ADMIN} Є обладнання з різними assignee.", tags="equipment,ui,filters",
           layer="UI",
           steps=[
               ("Відкрити /equipment", "Таблиця"),
               ("Відкрити dropdown «Відповідальний»", "Список співробітників"),
               ("Обрати employee A", "Таблиця лише з обладнанням A"),
           ]),
    ]


# --------------------------------------------------------------------------- #
# New Features / AC for import
# --------------------------------------------------------------------------- #

def extra_features() -> list[tuple]:
    return [
        (FEAT_CREW_ROOT, "REQ-WMS", "Екіпажі та видача на екіпаж",
         "Облік ресурсів на балансі екіпажів: області CREWS, видача UNIT→CREW, звіти та інвентаризація.",
         "WMS", "CRITICAL", "1", "7"),
        (FEAT_CREW_REG, FEAT_CREW_ROOT, "Області видимості екіпажів (CREWS)",
         "ADMIN керує StorageRegion accessMode=CREWS: locations (UNIT), members, hasCrews, crew-units API.",
         "WMS", "CRITICAL", "2", "0"),
        (FEAT_CREW_ISSUE, FEAT_CREW_ROOT, "Видача ресурсів на екіпаж",
         "OWNER видає на CREW/FLY_POINT (CREATED→FINISHED відправником); incident LOST; attached auto-forward; UI.",
         "WMS", "CRITICAL", "2", "1"),
        (FEAT_CREW_INV, FEAT_CREW_ROOT, "Залишки та інвентаризація екіпажів",
         "GET /inventory/crews (STOCK/INCOME); direct inventory; Crew-Manager RBAC; Fight write-off.",
         "WMS", "HIGH", "2", "2"),
        (FEAT_RVW, "REQ-RES", "Перегляд ресурсів (Resource Viewer)",
         "Роль Resource Viewer: зведення переміщень, фільтри journal, my-units без UNIT.",
         "RES", "HIGH", "1", "3"),
    ]


def extra_acceptance_criteria() -> list[tuple]:
    crew_reg = [
        ("AC-01", "@Admin створює область accessMode=CREWS"),
        ("AC-02", "Locations області містять UNIT для пошуку екіпажів"),
        ("AC-03", "Members — підрозділи-споживачі отримують доступ до екіпажів"),
        ("AC-04", "GET області CREWS за id повертає коректний recipientStorage"),
        ("AC-05", "hasCrews=true для member у області CREWS"),
        ("AC-06", "Область без members → hasCrews=false"),
        ("AC-07", "GET /storages/names/crew-units — ієрархія UNIT"),
        ("AC-08", "GET /storages/names/crews?parentId= — екіпажі під UNIT"),
        ("AC-09", "Рекурсивний пошук екіпажів у дочірніх UNIT"),
        ("AC-10", "getCrewUnits повертає дерево з батьківським UNIT"),
    ]
    crew_issue = [
        ("AC-01", "Видача UNIT→CREW → CREATED→FINISHED відправником; stock ±"),
        ("AC-02", "Після send — «В дорозі»; після FINISHED — sent history"),
        ("AC-03", "Негатив: amount > stock → HTTP 400"),
        ("AC-04", "Multi-resource send: credit після FINISHED"),
        ("AC-05", "Підрозділ поза CREWS region не може send на чужий crew"),
        ("AC-06", "PRODUCTION sender: CREATED→FINISHED на CREW"),
        ("AC-07", "Видача видима в received history після FINISHED"),
        ("AC-08", "Multi-item: кожна позиція ≤ stock"),
        ("AC-09", "UNIT→CREW прихований для ACCOUNTANT у GET /relocations"),
        ("AC-10", "totalRemovedResources на sender збільшується після send (CREATED)"),
        ("AC-11", "UI: кнопка «Видати на екіпаж» видима для member з CREWS"),
        ("AC-12", "UI: happy path — «В дорозі» → finish → «Видано»"),
        ("AC-13", "UI: «Всі локації» приховує кнопку видачі"),
        ("AC-14", "UI: каскад UNIT → combobox екіпажів"),
        ("AC-15", "UI: чужий owner не бачить екіпажі"),
        ("AC-16", "UI: qty > stock блокує submit"),
        ("AC-17", "UI: картка «Видано» в історії операцій"),
        ("AC-18", "Send на FLY_POINT → CREATED→FINISHED"),
        ("AC-19", "Attached CREW → auto-forward на FLY_POINT після FINISHED"),
        ("AC-20", "Пряме CREW→FLY_POINT → AUTO_FINISHED"),
        ("AC-21", "Надзвичайна подія на CREW → LOST без credit"),
        ("AC-22", "Надзвичайна подія на FLY_POINT → LOST без credit"),
        ("AC-23", "Отримувач CREW не може FINISHED — лише відправник"),
        ("AC-24", "Відправник скасовує CREATED (RETURNED) і відновлює stock"),
    ]
    crew_inv = [
        ("AC-01", "STOCK-звіт /inventory/crews показує залишок після видачі"),
        ("AC-02", "INCOME-звіт агрегує видачі за період"),
        ("AC-03", "STOCK-звіт узгоджений з direct GET /storages/{crew}/inventory"),
        ("AC-04", "OWNER без Crew-Manager — direct inventory → 403"),
        ("AC-05", "Crew-Manager читає direct crew inventory → 200"),
        ("AC-06", "Підрозділ поза CREWS — 403/404 на crew inventory"),
        ("AC-07", "Відкриття inventory session: ADMIN дозволено, чужий owner — ні"),
        ("AC-08", "PUT inventory змінює amount на crew storage"),
        ("AC-09", "Fight integration: write-off після sync (dev only)"),
        ("AC-10", "GET /write-off/short-stats — contract probe"),
        ("AC-11", "UI: залишки екіпажу на /unit-management"),
        ("AC-12", "UI: режим екіпажів без вибору crew — порожня таблиця"),
        ("AC-13", "UI: toggle інвентаризації після вибору crew"),
    ]
    rvw = [
        ("AC-01", "GET relocations viewer — sender лише STORAGE/PRODUCTION"),
        ("AC-02", "UNIT→CREW / UNIT→UNIT приховані у viewer journal"),
        ("AC-03", "GET relocations/sum — сортування resourceName ASC (digits→latin→ukr)"),
    ]
    rows: list[tuple] = []
    for i, (ac, text) in enumerate(crew_reg):
        rows.append((FEAT_CREW_REG, ac, text, str(i)))
    for i, (ac, text) in enumerate(crew_issue):
        rows.append((FEAT_CREW_ISSUE, ac, text, str(i)))
    for i, (ac, text) in enumerate(crew_inv):
        rows.append((FEAT_CREW_INV, ac, text, str(i)))
    for i, (ac, text) in enumerate(rvw):
        rows.append((FEAT_RVW, ac, text, str(i)))
    rows.append((FEAT_PLAN, "AC-03",
                 "Admin на /plans з «Всі локації»: сортування from DESC, потім storage.name ASC", "2"))
    return rows


def all_cases() -> list[Case]:
    builders = [
        accountant_cases,
        storage_crud_cases,
        storage_relation_inventory_cases,
        crew_region_cases,
        crew_relocation_cases,
        crew_inventory_cases,
        crew_ui_cases,
        relocation_journal_cases,
        relocation_invoice_ui_cases,
        visibility_region_case,
        resource_viewer_cases,
        tech_map_plan_guard_cases,
        global_plan_and_plans_cases,
        equipment_cases,
    ]
    seen: set[str] = set()
    unique: list[Case] = []
    for fn in builders:
        for case in fn():
            if case.test_id in seen:
                continue
            seen.add(case.test_id)
            unique.append(case)
    return unique


def filter_missing_from_export(cases: list[Case]) -> list[Case]:
    if not TCM_EXPORT.exists():
        return cases
    from openpyxl import load_workbook
    wb = load_workbook(TCM_EXPORT, read_only=True)
    ws = wb["TestCases"]
    existing = {str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}
    return [c for c in cases if c.test_id not in existing]


def main() -> None:
    cases = filter_missing_from_export(all_cases())
    write_xlsx_with_features(
        cases,
        OUTPUT,
        features=extra_features(),
        acceptance_criteria=extra_acceptance_criteria(),
        meta_extra=[("source", "erp-auto-test gap vs TCM 2026-07-02")],
    )
    print(f"Wrote {len(cases)} test cases to {OUTPUT}")
    domains: dict[str, int] = {}
    for case in cases:
        parts = case.test_id.split("-")
        key = parts[1] if len(parts) > 1 else "other"
        domains[key] = domains.get(key, 0) + 1
    for k in sorted(domains):
        print(f"  {k}: {domains[k]}")


if __name__ == "__main__":
    main()
