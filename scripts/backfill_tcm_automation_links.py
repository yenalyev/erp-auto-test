# -*- coding: utf-8 -*-
"""Backfill automation links on TCM cases via AI API PUT."""
from __future__ import annotations

import argparse
import json
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

BASE = "http://localhost:8100"
TOKEN = "dev-ai-token"
PROJECT_ID = 1
DEFAULT_XLSX = Path(r"D:\auto_test\tcm-import-gap-232-missing.xlsx")
DEFAULT_REPORT = Path(r"D:\auto_test\tcm-automation-backfill-report.txt")

HDR = {
    "X-TCM-Ai-Token": TOKEN,
    "Accept": "application/json",
    "Content-Type": "application/json; charset=utf-8",
}


def get_case(test_id: str) -> dict:
    url = f"{BASE}/api/ai/projects/{PROJECT_ID}/test-cases/{urllib.parse.quote(test_id, safe='')}"
    req = urllib.request.Request(url, headers=HDR)
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))


def put_case(test_id: str, body: dict) -> dict:
    url = f"{BASE}/api/ai/projects/{PROJECT_ID}/test-cases/{urllib.parse.quote(test_id, safe='')}"
    data = json.dumps(body, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=HDR, method="PUT")
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))


def load_links_from_xlsx(xlsx: Path) -> dict[str, dict[str, list[str]]]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["AutomationLinks"]
    rows = list(ws.iter_rows(values_only=True))[1:]
    by_tid: dict[str, dict[str, list[str]]] = defaultdict(lambda: {"API": [], "UI": []})
    for row in rows:
        if not row or not row[0] or not row[2]:
            continue
        test_id = str(row[0]).strip()
        layer = str(row[1] or "API").strip().upper()
        auto_id = str(row[2]).strip()
        if layer not in ("API", "UI"):
            layer = "API"
        by_tid[test_id][layer].append(auto_id)
    return by_tid


def load_links_from_list(list_path: Path) -> dict[str, dict[str, list[str]]]:
    """Self-link each testId: UI layer if TC-UI-*, else API. Skip cases that already have links."""
    by_tid: dict[str, dict[str, list[str]]] = {}
    for line in list_path.read_text(encoding="utf-8").splitlines():
        tid = line.strip()
        if not tid or tid.startswith("#"):
            continue
        if tid.upper().startswith("TC-UI-"):
            by_tid[tid] = {"API": [], "UI": [tid]}
        else:
            by_tid[tid] = {"API": [tid], "UI": []}
    return by_tid


def apply_links(by_tid: dict[str, dict[str, list[str]]], report: Path, *, skip_if_linked: bool) -> int:
    print(f"cases={len(by_tid)}")
    ok = fail = skipped = 0
    errors: list[str] = []

    for tid, layers in sorted(by_tid.items()):
        try:
            existing = get_case(tid)
            already_api = existing.get("apiAutomationIds") or []
            already_ui = existing.get("uiAutomationIds") or []
            if skip_if_linked and (already_api or already_ui):
                skipped += 1
                continue
            body = {
                "title": existing["title"],
                "apiAutomationIds": layers["API"],
                "uiAutomationIds": layers["UI"],
            }
            # When self-linking only one layer, keep the other layer as-is if present
            if skip_if_linked:
                if not layers["API"] and already_api:
                    body["apiAutomationIds"] = already_api
                if not layers["UI"] and already_ui:
                    body["uiAutomationIds"] = already_ui
            updated = put_case(tid, body)
            has = bool(updated.get("apiAutomationIds") or updated.get("uiAutomationIds"))
            if has or updated.get("automation"):
                ok += 1
            else:
                fail += 1
                errors.append(f"{tid}: no links after update")
        except urllib.error.HTTPError as e:
            fail += 1
            detail = e.read().decode("utf-8", errors="replace")
            errors.append(f"{tid}: HTTP {e.code} {detail}")
        except Exception as e:  # noqa: BLE001
            fail += 1
            errors.append(f"{tid}: {e}")

    lines = [
        f"updated_ok={ok}",
        f"failed={fail}",
        f"skipped_already_linked={skipped}",
        f"total={len(by_tid)}",
    ]
    if errors:
        lines.append("")
        lines.extend(errors[:80])
    report.parent.mkdir(parents=True, exist_ok=True)
    report.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print("\n".join(lines[:6]))
    print("wrote", report)
    return 0 if fail == 0 else 2


def main() -> int:
    parser = argparse.ArgumentParser(description="Backfill TCM automation links via AI API")
    parser.add_argument(
        "--from-list",
        type=Path,
        help="Text file with one testId per line (self-link; UI if TC-UI-*)",
    )
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="XLSX with AutomationLinks sheet")
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT, help="Report output path")
    args = parser.parse_args()

    if args.from_list:
        by_tid = load_links_from_list(args.from_list)
        report = args.report
        if report == DEFAULT_REPORT:
            report = Path(r"D:\auto_test\tcm-automation-backfill-138-report.txt")
        print(f"source=list {args.from_list}")
        return apply_links(by_tid, report, skip_if_linked=True)

    by_tid = load_links_from_xlsx(args.xlsx)
    print(f"source=xlsx {args.xlsx}")
    return apply_links(by_tid, args.report, skip_if_linked=False)


if __name__ == "__main__":
    raise SystemExit(main())
