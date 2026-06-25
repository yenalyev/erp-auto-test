#!/usr/bin/env python3
"""
Generate TCM import XLSX for manual QA — gaps vs export 2026-06-23.

Compare: python scripts/gap_tmp.py (228 automation @TestCaseId vs 228 TCM cases → 54 missing).

Includes Features + AcceptanceCriteria for REQ-GLOBAL-PLAN and integration ACs
(required for TCM import when featureId/acId pairs are absent in the project).
"""
from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from tcm_import_common import AUTHOR, ROLE_ADMIN, ROLE_OWNER, Case, Step

OUTPUT = Path(__file__).resolve().parent.parent / "docs" / "tcm-import-gap-automation-20260623.xlsx"

FEAT_GP = "REQ-GLOBAL-PLAN"
FEAT_INT = "REQ-INTEGRATION"
FEAT_INT_WA = "REQ-INTEGRATION-001"
FEAT_INT_DEL = "REQ-INTEGRATION-002"
FEAT_RES = "REQ-RES-001"
FEAT_RES_LIFE = "REQ-RES-002"
FEAT_RES_UI = "REQ-RES-003"
FEAT_NSP = "REQ-NON-SER-MAN"
FEAT_WMS_STOCK = "REQ-WMS-007"
FEAT_WMS_TRACK = "REQ-WMS-005"
FEAT_OPER_HIST = "REQ-OPER-HIST"
FEAT_RELOC = "REQ-EDIT_REL"
FEAT_AUTH = "REQ-RES-001"

ROLE_RVW = "ResourceViewer"
PRE_ADMIN = "@Admin залогінений. Середовище dev/staging."
PRE_OWNER = "@Owner 1 залогінений. У шапці SPA обрана конкретна локація Owner 1 (не «Всі локації»)."
PRE_GP = f"{PRE_ADMIN} Підготовлені техкарти M1/M2/M3 та ресурси A/B/C (див. docs/REQ-GLOBAL-PLAN-MAN.md)."
PRE_NSP = f"{PRE_OWNER} На складі є тестова сировина з відомим залишком."
PRE_BOT = "Середовище dev/staging. Налаштовані OAuth2 client_credentials для whatsapp-bot і delivery-bot."


def c(
    test_id: str,
    feature_id: str,
    ac_id: str,
    title: str,
    description: str,
    *,
    priority: str = "HIGH",
    severity: str = "MAJOR",
    preconditions: str = "",
    expected_result: str = "",
    tags: str = "",
    role_name: str = ROLE_ADMIN,
    layer: str | None = "API",
    steps: list[tuple[str, str]],
    cross: list[str] | None = None,
) -> Case:
    return Case(
        test_id=test_id,
        feature_id=feature_id,
        ac_id=ac_id,
        title=title,
        description=description,
        priority=priority,
        severity=severity,
        preconditions=preconditions,
        expected_result=expected_result or (steps[-1][1] if steps else ""),
        tags=tags,
        role_name=role_name,
        automation_layer=layer,
        automation_test_id=test_id,
        cross_features=cross or [],
        steps=[Step(i + 1, a, e) for i, (a, e) in enumerate(steps)],
    )


def bot_cases() -> list[Case]:
    return [
        c(
            "TC-BOT-001", FEAT_INT, "AC-INT-01",
            "OAuth2 client_credentials — отримання токена за SLA",
            "Ручна/інтеграційна перевірка: Keycloak видає access_token для ботів протягом ≤30 с.",
            priority="CRITICAL", preconditions=PRE_BOT, tags="integration,bots,sla",
            layer="API",
            steps=[
                ("POST на GET_TOKEN_URL з grant_type=client_credentials, client_id/secret whatsapp-bot або delivery-bot",
                 "HTTP 200; access_token не порожній; час відповіді ≤30 с"),
            ],
        ),
        c(
            "TC-BOT-002", FEAT_INT_WA, "AC-INT-01",
            "whatsapp-bot — GET internal/storages за SLA",
            "Ручна перевірка ендпоінту синхронізації складів для whatsapp-bot.",
            priority="CRITICAL", preconditions=PRE_BOT, tags="integration,bots,whatsapp",
            layer="API",
            steps=[
                ("Отримати Bearer token (client_credentials)", "Токен отримано"),
                ("GET /api/v1/internal/storages з Authorization: Bearer", "HTTP 200; JSON-масив складів; час ≤30 с"),
            ],
        ),
        c(
            "TC-BOT-003", FEAT_INT_DEL, "AC-INT-01",
            "delivery-bot — GET internal/relocations за SLA",
            "Ручна перевірка ендпоінту синхронізації переміщень для delivery-bot.",
            priority="CRITICAL", preconditions=PRE_BOT, tags="integration,bots,delivery",
            layer="API",
            steps=[
                ("Отримати Bearer token (client_credentials)", "Токен отримано"),
                ("GET /api/v1/internal/relocations з Authorization: Bearer", "HTTP 200; JSON-масив переміщень; час ≤30 с"),
            ],
        ),
    ]


def global_plan_cases() -> list[Case]:
    return [
        c("TC-GP-001", FEAT_GP, "AC-GP-01", "Список глобальних планів — фільтр month/year",
          "GET /api/v1/global-plans?month=&year= повертає створений план.",
          preconditions=PRE_GP, tags="global-plans,api",
          steps=[("Створити глобальний план на унікальний місяць", "План створено"),
                 ("GET /global-plans з month/year створеного плану", "У списку є id створеного плану")]),
        c("TC-GP-002", FEAT_GP, "AC-GP-01", "Створення глобального плану — happy path",
          "POST /api/v1/global-plans з output A=10 од.",
          preconditions=PRE_GP, tags="global-plans,api", priority="CRITICAL",
          steps=[("POST глобальний план: унікальний місяць, output ресурс A = 10", "HTTP 200; id, month, year, output[1] заповнені")]),
        c("TC-GP-003", FEAT_GP, "AC-GP-01", "Дублікат місяця — помилка",
          "Другий глобальний план на той самий календарний місяць відхиляється.",
          preconditions=PRE_GP, tags="global-plans,validation",
          steps=[("Створити план на місяць M", "Успіх"),
                 ("Повторити POST на той самий month/year", "HTTP 400; повідомлення про існуючий план")]),
        c("TC-GP-004", FEAT_GP, "AC-GP-02", "Output без техкарти — помилка",
          "Ресурс без активної PRODUCTION техкарти не може бути output.",
          preconditions=PRE_GP, tags="global-plans,validation",
          steps=[("POST з output-ресурсом без техкарти виробництва", "HTTP 400")]),
        c("TC-GP-005", FEAT_GP, "AC-GP-01", "Оновлення глобального плану (PUT)",
          "PUT змінює description/output; month/year без змін.",
          preconditions=PRE_GP, tags="global-plans,api",
          steps=[("Створити план", "id отримано"),
                 ("PUT з оновленим description (суфікс UPDATED)", "HTTP 200; period без змін; description оновлено")]),
        c("TC-GP-006", FEAT_GP, "AC-GP-09", "GET by id після generate — decomposition і generatedPlans",
          "Після генерації GET /global-plans/{id} містить знімок декомпозиції.",
          preconditions=PRE_GP, tags="global-plans,api",
          steps=[("Створити план і виконати generate (повна декомпозиція 3 блоки)", "Плани на локаціях створено"),
                 ("GET /global-plans/{id}", "generatedPlans не порожній; decomposition не null")]),
        c("TC-GP-007", FEAT_GP, "AC-GP-10", "DELETE глобального плану не видаляє location plans",
          "Видалення глобального плану зберігає згенеровані плани на локаціях.",
          preconditions=PRE_GP, tags="global-plans,api",
          steps=[("Generate location plans", "id планів L1 зафіксовано"),
                 ("DELETE /global-plans/{id}", "HTTP 2xx"),
                 ("GET /plans для L1", "Згенеровані location-плани все ще присутні")]),
        c("TC-GP-008", FEAT_GP, "AC-GP-07", "Експорт зведення потреб (XLSX)",
          "POST /global-plans/requirements/export повертає Excel.",
          preconditions=PRE_GP, tags="global-plans,export",
          steps=[("POST requirements/export з тестовими semiFinished/rawMaterials", "HTTP 200; Content-Type excel; файл не порожній")]),
        c("TC-GP-020", FEAT_GP, "AC-GP-04", "Декомпозиція блок 0 — echo з options і autoAssignable",
          "Перший POST /decompose повертає варіанти техкарт для ресурсу A.",
          preconditions=PRE_GP, tags="global-plans,decompose",
          steps=[("POST /decompose з emptyFirstBlock", "options не порожні; autoAssignable=true; requiredAmount=10")]),
        c("TC-GP-021", FEAT_GP, "AC-GP-05", "Декомпозиція — наступний блок потребує B=20",
          "Після A=10@L1/M1 система обчислює потребу B=20 (M1: 2B+3x→1A).",
          preconditions=PRE_GP, tags="global-plans,decompose",
          steps=[("POST decompose: блок A assignment 10@L1/M1", "complete=false; nextBlock містить B requiredAmount=20")]),
        c("TC-GP-022", FEAT_GP, "AC-GP-11", "Декомпозиція — кредит побічного продукту C",
          "Побічний C з M1 зменшує чисту потребу в C до 10 (не 20).",
          preconditions=PRE_GP, tags="global-plans,decompose",
          steps=[("Decompose блоки A і B (12@L1+8@L2)", "nextBlock: C requiredAmount=10")]),
        c("TC-GP-023", FEAT_GP, "AC-GP-06", "Повна декомпозиція — requirements і locationPlans",
          "3 блоки A→B→C повертають complete=true.",
          preconditions=PRE_GP, tags="global-plans,decompose", priority="CRITICAL",
          steps=[("POST decompose з повною декомпозицією", "complete=true; semiFinished не порожній; locationPlans є L1/L2")]),
        c("TC-GP-024", FEAT_GP, "AC-GP-08", "willReplace при існуючому плані локації",
          "Декомпозиція позначає L1 willReplace=true якщо план на місяць вже є.",
          preconditions=PRE_GP, tags="global-plans,decompose",
          steps=[("Створити location-план L1 на той самий місяць", "existingPlanId відомий"),
                 ("Повна декомпозиція", "locationPlans[L1].willReplace=true; existingPlanId збігається")]),
        c("TC-GP-025", FEAT_GP, "AC-GP-05", "Over-assignment — HTTP 400",
          "Сума assignment > output плану відхиляється.",
          preconditions=PRE_GP, tags="global-plans,validation",
          steps=[("Decompose з A=15 при output=10", "HTTP 400")]),
        c("TC-GP-026", FEAT_GP, "AC-GP-05", "Stale block — HTTP 400",
          "Пропущений блок A (лише B) відхиляється.",
          preconditions=PRE_GP, tags="global-plans,validation",
          steps=[("Decompose лише з блоком B", "HTTP 400 block mismatch")]),
        c("TC-GP-027", FEAT_GP, "AC-GP-11", "Sub-product на тій самій локації — net output L1",
          "Після повної декомпозиції L1 має ненульовий output.",
          preconditions=PRE_GP, tags="global-plans,decompose",
          steps=[("Повна декомпозиція", "locationPlans[L1].output не порожній")]),
        c("TC-GP-040", FEAT_GP, "AC-GP-09", "Generate — плани на локаціях з межами місяця",
          "POST /generate створює ≥2 location plans.",
          preconditions=PRE_GP, tags="global-plans,generate", priority="CRITICAL",
          steps=[("POST /generate з повною декомпозицією", "≥2 плани; L1: month/year як у global; output не порожній")]),
        c("TC-GP-041", FEAT_GP, "AC-GP-09", "Generate замінює існуючий місячний план",
          "replaced=true для L1; новий id ≠ старого.",
          preconditions=PRE_GP, tags="global-plans,generate",
          steps=[("Існуючий план L1 на місяць M", "id зафіксовано"),
                 ("Generate", "replaced=true; новий plan.id ≠ existing")]),
        c("TC-GP-042", FEAT_GP, "AC-GP-09", "Згенеровані плани видимі через GET /plans",
          "Список планів L1 містить id з generate.",
          preconditions=PRE_GP, tags="global-plans,generate",
          steps=[("Generate", "id з відповіді"),
                 ("GET /plans?storageId=L1", "Список містить згенерований id")]),
        c("TC-GP-043", FEAT_GP, "AC-GP-09", "Знімок декомпозиції зберігається після generate",
          "GET global plan: decomposition.blocks має 3 блоки.",
          preconditions=PRE_GP, tags="global-plans,generate",
          steps=[("Generate", "Успіх"),
                 ("GET /global-plans/{id}", "decomposition.blocks.size=3")]),
        c("TC-GP-044", FEAT_GP, "AC-GP-09", "Generate з неповною декомпозицією — HTTP 400",
          "emptyFirstBlock без assignments не дозволяє generate.",
          preconditions=PRE_GP, tags="global-plans,validation",
          steps=[("POST /generate з emptyFirstBlock", "HTTP 400")]),
        c("TC-GP-UI-SMOKE-001", FEAT_GP, "AC-GP-12", "UI smoke — список глобальних планів (Admin)",
          "Admin відкриває /global-plans.",
          preconditions=PRE_ADMIN, tags="global-plans,ui,smoke", layer="UI",
          steps=[("Відкрити /global-plans під Admin", "Заголовок «Глобальні плани» видимий")]),
        c("TC-GP-UI-SMOKE-002", FEAT_GP, "AC-GP-12", "UI smoke — wizard створення",
          "Перехід на wizard «Декомпозиція виробничого плану».",
          preconditions=PRE_ADMIN, tags="global-plans,ui,smoke", layer="UI",
          steps=[("На /global-plans натиснути «Створити план»", "Wizard «Декомпозиція виробничого плану» відкрито")]),
        c("TC-GP-UI-SMOKE-003", FEAT_GP, "AC-GP-12", "UI smoke — 4 вкладки wizard; пізні disabled",
          "На новому wizard вкладки 3–4 заблоковані.",
          preconditions=PRE_ADMIN, tags="global-plans,ui,smoke", layer="UI",
          steps=[("Відкрити wizard створення", "4 вкладки видимі"),
                 ("Перевірити стан вкладок 3 і 4", "Вкладки 3–4 disabled до розподілу")]),
        c("TC-GP-UI-HP-001", FEAT_GP, "AC-GP-09", "UI happy path — створення і generate через wizard",
          "Повний прохід Tab 1–4: output A=10, декомпозиція, «Створити плани по локаціях».",
          preconditions=PRE_GP, tags="global-plans,ui", priority="CRITICAL", layer="UI",
          steps=[
              ("Tab 1: період, output A=10, «Створити план»", "План створено; Tab 2 активна"),
              ("Tab 2: призначити B 12@L1 + 8@L2, «Розподілити по локаціях»", "Tab 3–4 розблоковані"),
              ("Tab 3: перегляд потреб, «Далі»", "Tab 4 активна"),
              ("Tab 4: «Створити плани по локаціях», «Готово»", "Badge «Створено»; плани на /plans для L1/L2"),
          ]),
    ]


def resource_deactivation_cases() -> list[Case]:
    pre = f"{PRE_ADMIN} Ресурс без залишків і зв'язків для happy path."
    return [
        c("TC-RES-010", FEAT_RES_LIFE, "AC-01", "Деактивація ресурсу — happy path",
          "DELETE /resources/{id} приховує ресурс зі словника, autocomplete (без archive), цін; видно на «Деактивовані».",
          preconditions=pre, tags="resources,deactivation", priority="CRITICAL",
          steps=[("Створити ізольований ресурс", "Ресурс у словнику"),
                 ("DELETE /resources/{id}", "HTTP 200; active=false"),
                 ("Перевірити словник, autocomplete, ціни, сторінку деактивованих", "Прихований скрізь крім archive/deactivated page")]),
        c("TC-RES-011", FEAT_RES_LIFE, "AC-01", "Реактивація ресурсу (unarchive)",
          "PUT /resources/unarchive/{id} повертає ресурс у активний словник.",
          preconditions=pre, tags="resources,deactivation",
          steps=[("Деактивувати ресурс", "active=false"),
                 ("PUT /resources/unarchive/{id}", "HTTP 200; ресурс знову в активному словнику")]),
        c("TC-RES-012", FEAT_RES, "AC-02", "Заборона деактивації — залишки на складі",
          "Ресурс з stock>0 на internal локації не деактивується.",
          preconditions=PRE_ADMIN, tags="resources,deactivation,validation",
          steps=[("Обрати ресурс з залишком на складі", "stock>0"),
                 ("DELETE /resources/{id}", "HTTP 400; повідомлення про склад; ресурс active")]),
        c("TC-RES-013", FEAT_RES, "AC-04", "Заборона деактивації — активна техкарта",
          "Ресурс у input/output активної техкарти не деактивується.",
          preconditions=PRE_ADMIN, tags="resources,deactivation,validation",
          steps=[("Створити активну техкарту з ресурсом R", "Техкарта active"),
                 ("DELETE /resources/{id} для R", "HTTP 400; згадка технологічних карт")]),
        c("TC-RES-014", FEAT_RES, "AC-03", "Заборона деактивації — переміщення «В дорозі»",
          "Ресурс у CREATED relocation не деактивується.",
          preconditions=PRE_ADMIN, tags="resources,deactivation,validation",
          steps=[("Створити видачу storage→storage (CREATED)", "Переміщення в дорозі"),
                 ("DELETE /resources/{id}", "HTTP 400; згадка переміщень")]),
        c("TC-RES-015", FEAT_RES, "AC-05", "Заборона деактивації — сповіщення про залишки",
          "Ресурс у stock alert не деактивується.",
          preconditions=PRE_ADMIN, tags="resources,deactivation,validation",
          steps=[("Створити alert для ресурсу R", "Alert існує"),
                 ("DELETE /resources/{id}", "HTTP 400; згадка сповіщень")]),
        c("TC-RES-016", FEAT_RES, "AC-04", "Виробництво з деактивованою складовою — DELETE/PUT 400",
          "Після деактивації складової техкарти запис виробництва не редагується/видаляється.",
          preconditions=PRE_ADMIN, tags="resources,deactivation,production",
          steps=[("Створити виробництво; деактивувати складову ресурсу", "Ресурс деактивовано"),
                 ("DELETE і PUT виробництва", "HTTP 400; повідомлення про деактивований ресурс")]),
        c("TC-RES-020", FEAT_RES, "AC-01", "Autocomplete — фільтр categoryIds",
          "GET /resources/autocomplete?categoryIds= повертає лише обрану категорію.",
          preconditions=PRE_ADMIN, tags="resources,autocomplete",
          steps=[("Створити ресурси в категоріях A і B зі спільним search", "2 ресурси"),
                 ("GET autocomplete з categoryIds=A", "Лише ресурс категорії A")]),
        c("TC-RES-021", FEAT_RES, "AC-01", "Autocomplete без categoryIds — всі категорії",
          "Без categoryIds повертаються ресурси всіх категорій за search.",
          preconditions=PRE_ADMIN, tags="resources,autocomplete",
          steps=[("GET autocomplete без categoryIds за спільним префіксом", "Обидва тестові ресурси в результаті")]),
    ]


def non_series_cases() -> list[Case]:
    return [
        c("TC-NSP-001", FEAT_NSP, "AC-01", "API — створення 1 од.; списання сировини",
          "POST /non-series-production: stock зменшується на usagePerUnit × amount.",
          preconditions=PRE_NSP, tags="non-series,api", role_name=ROLE_OWNER, priority="CRITICAL",
          steps=[("Зафіксувати stock сировини", "Базове значення"),
                 ("POST create: 1 од., usagePerUnit відомий", "HTTP 200; status IN_PROGRESS"),
                 ("Перевірити stock", "stock = before - usagePerUnit")]),
        c("TC-NSP-002", FEAT_NSP, "AC-01", "API — нестача сировини (негатив)",
          "usagePerUnit > stock відхиляється HTTP 400.",
          preconditions=PRE_NSP, tags="non-series,validation", role_name=ROLE_OWNER,
          steps=[("POST з usagePerUnit > поточного stock", "HTTP 400; stock без змін; запис не створено")]),
        c("TC-NSP-003", FEAT_NSP, "AC-01", "API — 2 од.; списання = usagePerUnit × 2",
          "Створення 2 од. коректно множить витрату.",
          preconditions=PRE_NSP, tags="non-series,api", role_name=ROLE_OWNER,
          steps=[("POST create amount=2", "HTTP 200"),
                 ("Перевірити stock", "stock = before - usagePerUnit×2")]),
        c("TC-NSP-004", FEAT_NSP, "AC-08", "API — GET /total збігається з сумою списку",
          "total з фільтрами productSearch/statuses = sum(amount) у list.",
          preconditions=PRE_NSP, tags="non-series,api", role_name=ROLE_OWNER,
          steps=[("Створити записи з різними статусами та productSearch", "2+ записи"),
                 ("GET list і GET /total з тими ж фільтрами", "total = сума amount у list")]),
        c("TC-UI-NSP-001", FEAT_NSP, "AC-01", "UI — створити 1 од. і завершити",
          "Owner створює запис «В роботі», переводить у «Завершено».",
          preconditions=PRE_NSP, tags="non-series,ui", role_name=ROLE_OWNER, layer="UI",
          steps=[("Відкрити /non-series-production, створити 1 од. з витратою сировини", "Запис «В роботі»"),
                 ("Змінити статус на «Завершено»", "Статус оновлено в журналі")]),
        c("TC-UI-NSP-002", FEAT_NSP, "AC-01", "UI — створити 2 од. і завершити",
          "Два одиниці продукту з коректним списанням.",
          preconditions=PRE_NSP, tags="non-series,ui", role_name=ROLE_OWNER, layer="UI",
          steps=[("Створити 2 од. несерійного виробництва", "Запис у журналі"),
                 ("Завершити запис", "Статус «Завершено»; stock зменшився на 2×usage")]),
        c("TC-UI-NSP-003", FEAT_NSP, "AC-08", "UI — відображення total у фільтрованому списку",
          "Сума «Всього» на UI збігається з API /total.",
          preconditions=PRE_NSP, tags="non-series,ui", role_name=ROLE_OWNER, layer="UI",
          steps=[("Застосувати productSearch фільтр", "Список відфільтровано"),
                 ("Порівняти total на UI з GET /total", "Значення збігаються")]),
        c("TC-UI-NSP-004", FEAT_NSP, "AC-06", "UI — створення без додавання сировини",
          "Можна зберегти запис без рядків resourceUsage (AC-06).",
          preconditions=PRE_NSP, tags="non-series,ui", role_name=ROLE_OWNER, layer="UI",
          steps=[("Створити запис лише з продуктом, без сировини", "Запис збережено"),
                 ("Перевірити stock", "Залишки без змін")]),
    ]


def misc_cases() -> list[Case]:
    return [
        c("TC-UI-003", FEAT_AUTH, "AC-01", "UI — логін Admin",
          "Admin входить через Keycloak і потрапляє на /production.",
          priority="CRITICAL", tags="auth,ui,smoke", layer="UI",
          steps=[("Відкрити SPA, увійти credentials Admin", "Редірект на /production; журнал завантажено")]),
        c("TC-UI-004", FEAT_AUTH, "AC-01", "UI — логін Owner 1",
          "Owner 1 входить і бачить журнал виробництва.",
          priority="CRITICAL", tags="auth,ui,smoke", role_name=ROLE_OWNER, layer="UI",
          steps=[("Увійти credentials Owner 1", "Редірект на /production")]),
        c("TC-UI-005", FEAT_AUTH, "AC-01", "UI — вихід Admin",
          "Після «Вийти» повернення на форму логіну Keycloak.",
          priority="CRITICAL", tags="auth,ui,smoke", layer="UI",
          steps=[("Увійти Admin", "Сесія активна"),
                 ("Sidebar → «Вийти»", "URL Keycloak/login; форма логіну")]),
        c("TC-UI-006", FEAT_AUTH, "AC-01", "UI — вихід Owner 1",
          "Logout Owner 1 повертає на login.",
          priority="CRITICAL", tags="auth,ui,smoke", role_name=ROLE_OWNER, layer="UI",
          steps=[("Увійти Owner 1", "Сесія активна"),
                 ("«Вийти»", "Форма логіну Keycloak")]),
        c("TC-UI-RES-AC-001", FEAT_RES_UI, "AC-01", "UI — autocomplete фільтр категорій на «Відстеження ресурсів»",
          "RESOURCE_VIEWER: фільтр категорії обмежує autocomplete ресурсів.",
          role_name=ROLE_RVW, tags="resources,ui,autocomplete", layer="UI",
          steps=[("Відкрити /resources-viewer/relocation", "Сторінка завантажена"),
                 ("Обрати категорію A у фільтрі", "Autocomplete показує лише ресурси A"),
                 ("«Очистити» фільтри", "Autocomplete знову показує ресурси A і B")]),
        c("TC-RVW-001", FEAT_WMS_TRACK, "AC-01", "GET relocations/sum — сортування resourceName ASC",
          "Порядок: цифри → латиниця → укр. л/є/і/ї (natural order).",
          role_name=ROLE_RVW, tags="resource-viewer,api",
          steps=[("Створити 8 тестових ресурсів (111_, aaa_, …, їжа_)", "id відомі"),
                 ("GET /resources-viewer/relocations/sum?resourceIds=…", "HTTP 200; імена відсортовані naturalOrder")]),
        c("TC-UI-HIST-DIS-001", FEAT_OPER_HIST, "AC-01", "UI — «Вироблено» після розбору з фактичним «Усього»",
          "Історія операцій відображає фактичне «Усього», а не розрахунок за техкартою.",
          preconditions=PRE_OWNER, tags="history,ui,disassemble", role_name=ROLE_OWNER, layer="UI",
          steps=[("Через API створити розбір з «Усього» ≠ розрахунку техкарти", "Розбір створено"),
                 ("Відкрити /history, картка «Вироблено»", "Delta = фактичне «Усього»")]),
        c("TC-WMS-007-008", FEAT_WMS_STOCK, "AC-07", "UI — «Експорт даних» лише для Admin",
          "/export-analytics: sidebar і експорт доступні Admin; Owner отримує помилку.",
          preconditions=PRE_ADMIN, tags="stock,ui,rbac", layer="UI",
          steps=[("Admin: sidebar «Експорт даних», відкрити /export-analytics", "Сторінка доступна"),
                 ("Owner 1: пункт прихований; прямий URL + експорт", "Toast помилки / 403")]),
        c("TC-REL-060b", FEAT_RELOC, "AC-06", "RBAC — Owner не може DELETE переміщення",
          "DELETE /relocations/{id} під Owner 1 для external receive → 401/403.",
          preconditions=PRE_OWNER, tags="relocations,rbac,negative", role_name=ROLE_OWNER,
          steps=[("Створити зовнішнє отримання (Owner)", "Запис створено"),
                 ("Спроба DELETE під Owner 1", "HTTP 401 або 403")]),
    ]


def extra_features() -> list[tuple]:
    """Features absent from TCM export 2026-06-23."""
    return [
        (FEAT_GP, "REQ-MFG", "Глобальні плани (декомпозиція)",
         "Декомпозиція виробничого плану: /global-plans, wizard 4 вкладки, generate location plans.",
         "MFG", "HIGH", "0", "4"),
    ]


def extra_acceptance_criteria() -> list[tuple]:
    gp_ac = [
        ("AC-GP-01", "Один глобальний план на календарний місяць; CRUD output"),
        ("AC-GP-02", "Output — лише ресурси з активною PRODUCTION техкартою"),
        ("AC-GP-03", "Зміна Tab1 скидає декомпозицію"),
        ("AC-GP-04", "Одна карта+локація → auto-assign"),
        ("AC-GP-05", "Кілька карт/локацій → діалог призначення; валідація over-assignment"),
        ("AC-GP-06", "«Розподілити по локаціях» активує Tab 3/4"),
        ("AC-GP-07", "Tab3: напівфабрикати + сировина + залишки; експорт requirements"),
        ("AC-GP-08", "Tab4: «Замінить наявний» при існуючому плані локації"),
        ("AC-GP-09", "Generate створює Plan per storage"),
        ("AC-GP-10", "DELETE global plan не видаляє location plans"),
        ("AC-GP-11", "Проміжні компоненти на локації споживання (by-product credit)"),
        ("AC-GP-12", "RBAC global-plan::*; sidebar для Admin"),
    ]
    rows = [(FEAT_GP, ac, text, str(i)) for i, (ac, text) in enumerate(gp_ac)]
    rows.append((FEAT_INT, "AC-INT-01", "Internal API для ботів відповідає за SLA ≤30 с", "0"))
    rows.append((FEAT_INT_WA, "AC-INT-01", "GET /internal/storages для whatsapp-bot — HTTP 200", "0"))
    rows.append((FEAT_INT_DEL, "AC-INT-01", "GET /internal/relocations для delivery-bot — HTTP 200", "0"))
    rows.append((FEAT_RES_LIFE, "AC-01", "@Admin може деактивувати/реактивувати ресурс без зв'язків", "0"))
    rows.append((FEAT_WMS_TRACK, "AC-01", "Відстеження ресурсів: API sum і сортування за назвою", "0"))
    return rows


def all_cases() -> list[Case]:
    cases: list[Case] = []
    cases.extend(bot_cases())
    cases.extend(global_plan_cases())
    cases.extend(resource_deactivation_cases())
    cases.extend(non_series_cases())
    cases.extend(misc_cases())
    seen: set[str] = set()
    unique: list[Case] = []
    for case in cases:
        if case.test_id in seen:
            continue
        seen.add(case.test_id)
        unique.append(case)
    return unique


def write_xlsx_with_features(cases: list[Case], path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    meta = wb.create_sheet("Meta")
    meta.append(["key", "value"])
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")
    for row in [
        ("formatVersion", "1"),
        ("exportedAt", now),
        ("projectId", "1"),
        ("projectName", "ERP Система"),
        ("scope", "PROJECT"),
        ("rootFeatureId", ""),
        ("source", "erp-auto-test gap vs TCM 2026-06-23"),
    ]:
        meta.append(list(row))

    feat_sheet = wb.create_sheet("Features")
    feat_sheet.append([
        "featureId", "parentFeatureId", "title", "description", "module",
        "priority", "status", "author", "treeDepth", "sortOrder",
    ])
    for row in extra_features():
        feat_sheet.append([*row, "ACTIVE", AUTHOR])

    ac_sheet = wb.create_sheet("AcceptanceCriteria")
    ac_sheet.append(["featureId", "acId", "text", "sortOrder"])
    for row in extra_acceptance_criteria():
        ac_sheet.append(list(row))

    tc_sheet = wb.create_sheet("TestCases")
    tc_sheet.append([
        "testId", "featureId", "acId", "title", "description", "priority", "severity", "status",
        "testType", "preconditions", "expectedResult", "tags", "author", "jiraIssueKey", "roleName",
        "parameterized", "dependencies",
    ])
    for case in cases:
        tc_sheet.append([
            case.test_id, case.feature_id, case.ac_id, case.title, case.description,
            case.priority, case.severity, "ACTIVE", case.test_type,
            case.preconditions, case.expected_result, case.tags, AUTHOR, "", case.role_name,
            "false", "",
        ])

    steps_sheet = wb.create_sheet("Steps")
    steps_sheet.append(["testId", "stepOrder", "actionText", "expectedText"])
    for case in cases:
        for step in case.steps:
            steps_sheet.append([case.test_id, str(step.order), step.action, step.expected])

    schema_sheet = wb.create_sheet("DatasetSchema")
    schema_sheet.append(["testId", "fieldKey", "fieldLabel", "fieldType", "required", "sortOrder"])

    params_sheet = wb.create_sheet("ParameterSets")
    params_sheet.append(["testId", "setName", "active", "valuesJson"])

    auto_sheet = wb.create_sheet("AutomationLinks")
    auto_sheet.append(["testId", "layer", "automationTestId", "sortOrder"])
    for case in cases:
        if case.automation_layer and case.automation_test_id:
            auto_sheet.append([case.test_id, case.automation_layer, case.automation_test_id, "0"])

    cross_sheet = wb.create_sheet("CrossFeatures")
    cross_sheet.append(["testId", "crossFeatureSlug"])
    for case in cases:
        for slug in case.cross_features:
            cross_sheet.append([case.test_id, slug])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    cases = all_cases()
    write_xlsx_with_features(cases, OUTPUT)
    print(f"Wrote {len(cases)} manual test cases to {OUTPUT}")
    domains = {}
    for case in cases:
        prefix = case.test_id.split("-")[1] if "-" in case.test_id else "other"
        domains[prefix] = domains.get(prefix, 0) + 1
    for k in sorted(domains):
        print(f"  {k}: {domains[k]}")


if __name__ == "__main__":
    main()
