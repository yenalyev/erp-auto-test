#!/usr/bin/env python3
"""
Generate TCM test-case import XLSX for manual QA (format TestCaseXlsxIO v1).

Gaps: automation @TestCaseId not present in TCM export 2026-06-18.
Only uses featureId/acId pairs that exist in that export.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook

OUTPUT = Path(__file__).resolve().parent.parent / "docs" / "tcm-import-erp-manual-20260618.xlsx"

AUTHOR = "QA"
ROLE_OWNER = "Owner"
ROLE_ADMIN = "Admin"


@dataclass
class Step:
    order: int
    action: str
    expected: str


@dataclass
class Case:
    test_id: str
    feature_id: str
    ac_id: str
    title: str
    description: str
    priority: str = "HIGH"
    severity: str = "MAJOR"
    test_type: str = "FUNCTIONAL"
    preconditions: str = ""
    expected_result: str = ""
    tags: str = ""
    role_name: str = ROLE_OWNER
    steps: list[Step] = field(default_factory=list)
    automation_layer: str | None = None  # API | UI
    automation_test_id: str | None = None
    cross_features: list[str] = field(default_factory=list)


def prd_cases() -> list[Case]:
    base_pre = "@Owner залогінений. Обрана локація Owner 1. На складі достатньо сировини для техкарти тесту."
    cases = [
        Case(
            "TC-PRD-001", "REQ-MFG-002", "AC-01",
            "Створення виробництва — happy path",
            "Ручна перевірка сценарію створення партії виробництва та впливу на залишки.",
            "CRITICAL", "MAJOR", preconditions=base_pre,
            expected_result="Запис у журналі; сировина списана; готова продукція на складі.",
            tags="production,api-regression",
            automation_layer="API", automation_test_id="TC-PRD-001",
            cross_features=["batch", "account"],
            steps=[
                Step(1, "Створити виробництво через UI «Виготовлення» або POST /api/v1/productions/{storageId}",
                     "HTTP 200; партія з'являється в журналі /production"),
                Step(2, "Перевірити залишки сировини та готової продукції на складі",
                     "Сировина зменшилась; продукція збільшилась відповідно до техкарти"),
            ],
        ),
        Case(
            "TC-PRD-002", "REQ-MFG-002", "AC-01",
            "Журнал виробництва — +1 запис після створення",
            "Перевірка, що після створення збільшується totalElements журналу.",
            preconditions=base_pre, tags="production,api-regression",
            automation_layer="API", automation_test_id="TC-PRD-002",
            steps=[Step(1, "Зафіксувати totalElements GET /api/v1/productions?storageIds={id}",
                        "Отримано базове значення N"),
                   Step(2, "Створити виробництво і повторити GET",
                        "totalElements = N + 1")],
        ),
        Case(
            "TC-PRD-003", "REQ-MFG-002", "AC-02",
            "Валідація: обсяг виробництва ≤ 0",
            "Негативні значення обсягу відхиляються.",
            preconditions=base_pre, tags="production,validation",
            automation_layer="API", automation_test_id="TC-PRD-003",
            steps=[Step(1, "Спробувати створити виробництво з amount=0 та amount<0",
                        "HTTP 400; запис не створено")],
        ),
        Case(
            "TC-PRD-004", "REQ-MFG-002", "AC-03",
            "Валідація: недостатня сировина",
            "Створення блокується при нестачі сировини.",
            preconditions=base_pre, tags="production,validation",
            automation_layer="API", automation_test_id="TC-PRD-004",
            cross_features=["batch"],
            steps=[Step(1, "Спробувати виробництво з обсягом, що перевищує залишок сировини",
                        "HTTP 400; залишки без змін")],
        ),
        Case(
            "TC-PRD-005", "REQ-MFG-002", "AC-04",
            "Дублікат номера партії",
            "Повторна партія з тим самим batchNumber відхиляється.",
            preconditions=base_pre, tags="production,validation",
            automation_layer="API", automation_test_id="TC-PRD-005",
            steps=[Step(1, "Створити виробництво з унікальною партією", "Успіх"),
                   Step(2, "Повторити створення з тією ж партією", "HTTP 400")],
        ),
        Case(
            "TC-PRD-006", "REQ-MFG-002", "AC-05",
            "Оновлення виробництва (зміна обсягу)",
            "PUT оновлює запис і перераховує залишки.",
            preconditions=base_pre, tags="production",
            automation_layer="API", automation_test_id="TC-PRD-006",
            cross_features=["batch"],
            steps=[Step(1, "Створити виробництво", "Запис у журналі"),
                   Step(2, "Змінити amount через PUT", "Новий обсяг у журналі; залишки оновлені")],
        ),
        Case(
            "TC-PRD-007", "REQ-MFG-002", "AC-06",
            "Видалення виробництва",
            "DELETE відкочує залишки та прибирає запис з журналу.",
            preconditions=base_pre, tags="production",
            automation_layer="API", automation_test_id="TC-PRD-007",
            steps=[Step(1, "Створити і видалити виробництво", "Запис зник; залишки як до створення")],
        ),
        Case(
            "TC-PRD-008", "REQ-MFG-002", "AC-06",
            "Дублікат партії при повторному створенні (контроль цілісності)",
            "Другий запис з існуючою партією не створюється.",
            preconditions=base_pre, tags="production,validation",
            automation_layer="API", automation_test_id="TC-PRD-008",
            steps=[Step(1, "Двічі надіслати create з однаковим batchNumber", "Лише перший успішний")],
        ),
    ]
    filters = [
        ("TC-PRD-FLT-001", "product", "Фільтр журналу за продуктом", "GET ?product=… містить опорний запис"),
        ("TC-PRD-FLT-002", "startDate", "Фільтр «З» (startDate)", "Усі записи ≥ обраної дати"),
        ("TC-PRD-FLT-003", "endDate", "Фільтр «По» (endDate)", "Усі записи ≤ обраної дати"),
        ("TC-PRD-FLT-004", "dateRange", "Діапазон дат", "Записи в межах діапазону"),
        ("TC-PRD-FLT-005", "category", "Фільтр за категорією", "Продукти лише обраної категорії"),
        ("TC-PRD-FLT-006", "product+dates", "Продукт + діапазон дат", "AND усіх умов"),
        ("TC-PRD-FLT-007", "product+category", "Продукт + категорія", "AND усіх умов"),
        ("TC-PRD-FLT-008", "all", "Продукт + дати + категорія", "Комбінований фільтр"),
    ]
    for tid, kind, title, exp in filters:
        cases.append(Case(
            tid, "REQ-MFG-002", "AC-01",
            f"Журнал виробництва — {title}",
            f"Ручна/API перевірка фільтра {kind} журналу (дублює автотест {tid}).",
            "MEDIUM", "MINOR", preconditions=base_pre, tags="production,filters",
            expected_result=exp,
            automation_layer="API", automation_test_id=tid,
            steps=[
                Step(1, f"Відкрити /production і застосувати фільтр ({kind}) або виконати GET /productions",
                     exp),
                Step(2, "Переконатися, що опорний запис з baseline присутній у результаті",
                     "Опорний id видно в таблиці або API content"),
            ],
        ))
    return cases


def ui_production_cases() -> list[Case]:
    pre = "@Owner залогінений. Обрана локація Owner 1 (не «Всі локації»)."
    cases = [
        Case(
            "TC-UI-PROD-001", "REQ-MFG-002", "AC-01",
            "UI smoke — журнал виробництва відображає записи",
            "Перевірка структури сторінки /production та наявності даних у таблиці.",
            "CRITICAL", "MAJOR", preconditions=pre, tags="production,ui,smoke",
            expected_result="Кнопки Виготовлення/Розбір, фільтри, таблиця з записами (не empty state).",
            automation_layer="UI", automation_test_id="TC-UI-PROD-001",
            steps=[
                Step(1, "Відкрити /production", "Сторінка завантажена без помилки"),
                Step(2, "Перевірити таблицю журналу", "Є рядки; опорний запис (перший у API) видно за партією"),
            ],
        ),
    ]
    ui_filters = [
        ("TC-UI-PROD-FLT-001", "продукт"),
        ("TC-UI-PROD-FLT-002", "дата «З»"),
        ("TC-UI-PROD-FLT-003", "дата «По»"),
        ("TC-UI-PROD-FLT-004", "діапазон дат"),
        ("TC-UI-PROD-FLT-005", "категорія"),
        ("TC-UI-PROD-FLT-006", "продукт + дати"),
        ("TC-UI-PROD-FLT-007", "продукт + категорія"),
        ("TC-UI-PROD-FLT-008", "продукт + дати + категорія"),
    ]
    for tid, label in ui_filters:
        cases.append(Case(
            tid, "REQ-MFG-002", "AC-01",
            f"UI — фільтр журналу: {label}",
            f"Ручна перевірка фільтра «{label}» на /production зі звіркою з API.",
            "HIGH", "MAJOR", preconditions=pre, tags="production,ui,filters",
            automation_layer="UI", automation_test_id=tid,
            steps=[
                Step(1, f"Застосувати фільтр «{label}» на UI", "Таблиця оновилась"),
                Step(2, "Знайти опорний запис за номером партії", "Рядок присутній; ключові поля збігаються"),
            ],
        ))
    return cases


def relocation_ui_cases() -> list[Case]:
    pre = "@Owner 1 залогінений. Обрана конкретна локація (не «Всі локації»). Є ресурс із залишком ≥ 50."
    mapping = [
        ("TC-UI-REL-001", "AC-03", "Зовнішнє отримання (постачальник → склад)", "relocations,ui"),
        ("TC-UI-REL-002", "AC-01", "Видача storage → storage та підтвердження", "relocations,ui"),
        ("TC-UI-REL-003", "AC-01", "Видача на UNIT (AUTO_FINISHED)", "relocations,ui"),
        ("TC-UI-REL-004", "AC-04", "Редагування зовнішнього отримання (зменшити кількість)", "relocations,ui"),
        ("TC-UI-REL-005", "AC-04", "Admin — редагування/видалення зовнішнього отримання", "relocations,ui", ROLE_ADMIN),
        ("TC-UI-REL-006", "AC-02", "Відхилення переміщення (RETURNED)", "relocations,ui"),
        ("TC-UI-REL-007", "AC-01", "Редагування видачі (зменшити кількість)", "relocations,ui"),
        ("TC-UI-REL-008", "AC-03", "Видача з явною партією", "relocations,ui,batch"),
    ]
    cases = []
    for item in mapping:
        tid, ac, title, tags = item[0], item[1], item[2], item[3]
        role = item[4] if len(item) > 4 else ROLE_OWNER
        cases.append(Case(
            tid, "REQ-EDIT_REL", ac, f"UI — {title}",
            f"Ручне проходження сценарію переміщення ({title}). Деталі: docs/REQ-RELOCATION-MAN.md",
            "HIGH", "MAJOR", preconditions=pre, tags=tags, role_name=role,
            automation_layer="UI", automation_test_id=tid,
            cross_features=["relocation", "batch"] if "batch" in tags else ["relocation"],
            steps=[
                Step(1, "Відкрити /relocations і виконати сценарій згідно REQ-RELOCATION-MAN", "Операція успішна на UI"),
                Step(2, "Перевірити статус у журналі та залишки/партії", "Відповідає AC з документації"),
            ],
        ))
    for tid, title in [
        ("TC-UI-REL-EQ-001", "Вкладка обладнання — smoke журналу"),
        ("TC-UI-REL-EQ-002", "Обладнання — підтвердження FINISHED на UI"),
        ("TC-UI-REL-EQ-003", "Обладнання — removeInvoiceFile на UI"),
    ]:
        cases.append(Case(
            tid, "REQ-EQU-001", "AC-01", title,
            "Ручна перевірка UI переміщень обладнання.",
            "HIGH", "MAJOR", preconditions=pre, tags="equipment,relocations,ui",
            automation_layer="UI", automation_test_id=tid,
            cross_features=["equipment", "relocation"],
            steps=[Step(1, "Виконати сценарій на /relocations (обладнання)", "UI відповідає API")],
        ))
    return cases


def relocation_api_spot_cases() -> list[Case]:
    """Key API flows for manual spot-check (automation exists; TCM missing)."""
    pre = "@Owner 1. Достатній залишок на складі."
    specs = [
        ("TC-REL-010", "AC-01", "Видача storage → storage (CREATED)"),
        ("TC-REL-020", "AC-03", "Отримання SUPPLIER → storage"),
        ("TC-REL-030", "AC-02", "Підтвердження CREATED → FINISHED"),
        ("TC-REL-031", "AC-02", "Відхилення CANCELLED → RETURNED"),
        ("TC-REL-040", "AC-01", "Редагування AUTO_FINISHED send (зменшити qty)"),
        ("TC-REL-050", "AC-04", "Редагування receive (зменшити qty)"),
        ("TC-REL-057", "AC-04", "Admin видаляє зовнішнє отримання"),
        ("TC-REL-014", "AC-01", "Негатив: недостатній залишок при видачі"),
    ]
    return [
        Case(
            tid, "REQ-EDIT_REL", ac, title,
            f"API/ручна регресія: {title}. Див. docs/REQ-RELOCATION-MAN.md",
            "MEDIUM", "MAJOR", preconditions=pre, tags="relocations,api-regression",
            automation_layer="API", automation_test_id=tid,
            cross_features=["relocation", "batch"],
            steps=[Step(1, "Виконати кроки з REQ-RELOCATION-MAN для " + tid, "Очікуваний статус і залишки")],
        )
        for tid, ac, title in specs
    ]


def nsp_manual_gaps() -> list[Case]:
    pre = "@Owner залогінений. Обрана локація owner."
    return [
        Case(
            "TC-NON-SER-MAN-010", "REQ-NON-SER-MAN", "AC-02",
            "Owner не може редагувати «Завершено» після 2 днів",
            "Ручна перевірка обмеження 2 дні для Owner (AC-02). Автотестів немає.",
            "HIGH", "MAJOR", preconditions=pre + " Є запис «Завершено» старше 2 днів (або змінити дату в БД на dev).",
            tags="non-series,manual-only",
            steps=[
                Step(1, "Відкрити запис несерійного виробництва у статусі «Завершено» старше 2 днів",
                     "Кнопки редагування/видалення недоступні або операція відхилена"),
                Step(2, "Повторити під @Admin", "Admin може редагувати (AC-03)"),
            ],
        ),
        Case(
            "TC-NSP-005", "REQ-NON-SER-MAN", "AC-04",
            "Owner редагує «В роботі» без обмеження часу",
            "Ручна перевірка AC-04 для статусу «В роботі».",
            preconditions=pre, tags="non-series",
            automation_layer="API", automation_test_id="TC-NSP-005",
            steps=[Step(1, "Редагувати запис «В роботі»", "Зміни збережено")],
        ),
    ]


def dictionary_cases() -> list[Case]:
    pre = "@Admin залогінений."
    return [
        Case("TC-RES-002", "REQ-RES-001", "AC-01", "Ресурс — оновлення назви",
             "CRUD довідника ресурсів: PUT name.", preconditions=pre, tags="dictionary",
             automation_layer="API", automation_test_id="TC-RES-002",
             steps=[Step(1, "Оновити назву ресурсу через UI або API", "200; нова назва в списку")]),
        Case("TC-RES-003", "REQ-RES-001", "AC-01", "Ресурс — дублікат назви (негатив)",
             "Створення дубліката відхиляється.", preconditions=pre, tags="dictionary,validation",
             automation_layer="API", automation_test_id="TC-RES-003",
             steps=[Step(1, "POST ресурс з існуючою назвою", "HTTP 400")]),
        Case("TC-STR-001", "REQ-STR", "AC-01", "Склад — створення",
             "CRUD локацій.", preconditions=pre, tags="dictionary",
             automation_layer="API", automation_test_id="TC-STR-001",
             steps=[Step(1, "Створити STORAGE через UI/API", "Локація в списку")]),
        Case("TC-STR-002", "REQ-STR", "AC-01", "Локація — оновлення",
             "Оновлення всіх полів PUT; relation без змін.",
             preconditions=pre, tags="dictionary", automation_layer="API", automation_test_id="TC-STR-002",
             steps=[Step(1, "PUT оновити локацію", "200; зміни в GET by id")]),
        Case("TC-STR-003", "REQ-STR", "AC-01", "Локація — дублікат назви (негатив)",
             "Створення дубліката серед активних локацій відхиляється.",
             preconditions=pre, tags="dictionary,validation", automation_layer="API", automation_test_id="TC-STR-003",
             steps=[Step(1, "POST з існуючою назвою", "HTTP 400")]),
        Case("TC-STR-004", "REQ-STR", "AC-01", "Локація — валідація name (негатив)",
             "Порожня або відсутня назва відхиляється.",
             preconditions=pre, tags="dictionary,validation", automation_layer="API", automation_test_id="TC-STR-004",
             steps=[Step(1, "POST без name", "HTTP 400")]),
        Case("TC-STR-005", "REQ-STR", "AC-01", "Локація — relation INTERNAL→EXTERNAL (негатив)",
             "Заборона зміни relation з INTERNAL на EXTERNAL.",
             preconditions=pre, tags="dictionary,validation", automation_layer="API", automation_test_id="TC-STR-005",
             steps=[Step(1, "PUT з relation=EXTERNAL", "HTTP 400")]),
        Case("TC-STR-006", "REQ-STR", "AC-01", "Локація — архівація",
             "DELETE встановлює active=false; зникає з isActive=true names.",
             preconditions=pre, tags="dictionary", automation_layer="API", automation_test_id="TC-STR-006",
             steps=[Step(1, "DELETE /storages/{id}", "200; active=false")]),
        Case("TC-STR-007", "REQ-STR", "AC-01", "Локація — розархівація",
             "PUT unarchive повертає active=true.",
             preconditions=pre, tags="dictionary", automation_layer="API", automation_test_id="TC-STR-007",
             steps=[Step(1, "PUT /storages/unarchive/{id}", "200; active=true")]),
        Case("TC-STR-008", "REQ-STR", "AC-01", "Локація — повторне використання назви після архівації",
             "Унікальність name лише серед active; архівна назва доступна для нового POST.",
             preconditions=pre, tags="dictionary", automation_layer="API", automation_test_id="TC-STR-008",
             steps=[Step(1, "POST з назвою архівованої локації", "200")]),
        Case("TC-MU-001", "REQ-DICT-001", "AC-03", "Одиниця виміру — створення",
             "CRUD одиниць виміру: POST.",
             preconditions=pre, tags="dictionary", automation_layer="API", automation_test_id="TC-MU-001",
             steps=[Step(1, "Створити одиницю виміру", "201/200")]),
        Case("TC-MU-002", "REQ-DICT-001", "AC-03", "Одиниця виміру — оновлення",
             "PUT одиниця виміру.",
             preconditions=pre, tags="dictionary", automation_layer="API", automation_test_id="TC-MU-002",
             steps=[Step(1, "Оновити shortName", "200")]),
        Case("TC-MU-003", "REQ-DICT-001", "AC-03", "Одиниця виміру — дублікат (негатив)",
             "Дублікат shortName відхиляється.",
             preconditions=pre, tags="dictionary,validation", automation_layer="API", automation_test_id="TC-MU-003",
             steps=[Step(1, "Створити дублікат", "400")]),
    ]


def auth_ui_cases() -> list[Case]:
    return [
        Case("TC-UI-001", "REQ-RES-001", "AC-01", "UI — успішний логін",
             "Форма Keycloak → редірект у SPA.", "CRITICAL", "MAJOR",
             tags="auth,ui,smoke", automation_layer="UI", automation_test_id="TC-UI-001",
             steps=[Step(1, "Увійти валідними credentials", "Редірект на /production або першу дозволену сторінку")]),
        Case("TC-UI-002", "REQ-RES-001", "AC-01", "UI — невалідний пароль",
             "Негативний сценарій авторизації через Keycloak.",
             tags="auth,ui,negative", automation_layer="UI", automation_test_id="TC-UI-002",
             steps=[Step(1, "Спроба логіну з невірним паролем", "Залишаємось на формі логіну / помилка")]),
        Case("TC-SMOKE-001", "REQ-RES-001", "AC-01", "Smoke — health / доступність SPA",
             "Перевірка доступності фронтенду.",
             tags="smoke", automation_layer="API", automation_test_id="TC-SMOKE-001",
             steps=[Step(1, "Відкрити base URL", "SPA завантажується")]),
        Case("TC-SMOKE-002", "REQ-RES-001", "AC-01", "Smoke — захищений API без токена",
             "API без сесії повертає 401.",
             tags="smoke", automation_layer="API", automation_test_id="TC-SMOKE-002",
             steps=[Step(1, "GET /api/v1/resources без auth", "401")]),
    ]


def all_cases() -> list[Case]:
    cases: list[Case] = []
    cases.extend(prd_cases())
    cases.extend(ui_production_cases())
    cases.extend(relocation_ui_cases())
    cases.extend(relocation_api_spot_cases())
    cases.extend(nsp_manual_gaps())
    cases.extend(dictionary_cases())
    cases.extend(auth_ui_cases())
    # de-duplicate by test_id
    seen: set[str] = set()
    unique: list[Case] = []
    for c in cases:
        if c.test_id in seen:
            continue
        seen.add(c.test_id)
        unique.append(c)
    return unique


def write_xlsx(cases: list[Case], path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    meta = wb.create_sheet("Meta")
    meta.append(["key", "value"])
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")
    for row in [
        ("formatVersion", "1"),
        ("exportedAt", now),
        ("projectId", "1"),
        ("projectName", "ERP Система"),
        ("scope", "PROJECT"),
        ("rootFeatureId", ""),
    ]:
        meta.append(list(row))

    tc_sheet = wb.create_sheet("TestCases")
    tc_sheet.append([
        "testId", "featureId", "acId", "title", "description", "priority", "severity", "status",
        "testType", "preconditions", "expectedResult", "tags", "author", "jiraIssueKey", "roleName",
        "parameterized", "dependencies",
    ])
    for c in cases:
        tc_sheet.append([
            c.test_id, c.feature_id, c.ac_id, c.title, c.description,
            c.priority, c.severity, "ACTIVE", c.test_type,
            c.preconditions, c.expected_result, c.tags, AUTHOR, "", c.role_name,
            "false", "",
        ])

    steps_sheet = wb.create_sheet("Steps")
    steps_sheet.append(["testId", "stepOrder", "actionText", "expectedText"])
    for c in cases:
        for s in c.steps:
            steps_sheet.append([c.test_id, str(s.order), s.action, s.expected])

    schema_sheet = wb.create_sheet("DatasetSchema")
    schema_sheet.append(["testId", "fieldKey", "fieldLabel", "fieldType", "required", "sortOrder"])

    params_sheet = wb.create_sheet("ParameterSets")
    params_sheet.append(["testId", "setName", "active", "valuesJson"])

    auto_sheet = wb.create_sheet("AutomationLinks")
    auto_sheet.append(["testId", "layer", "automationTestId", "sortOrder"])
    for c in cases:
        if c.automation_layer and c.automation_test_id:
            auto_sheet.append([c.test_id, c.automation_layer, c.automation_test_id, "0"])

    cross_sheet = wb.create_sheet("CrossFeatures")
    cross_sheet.append(["testId", "crossFeatureSlug"])
    for c in cases:
        for slug in c.cross_features:
            cross_sheet.append([c.test_id, slug])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {len(cases)} test cases to {path}")


if __name__ == "__main__":
    write_xlsx(all_cases(), OUTPUT)
