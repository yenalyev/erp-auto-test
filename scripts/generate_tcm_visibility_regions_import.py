#!/usr/bin/env python3
"""
Generate TCM import XLSX for visibility regions (locations + resources).

Features: REQ-REGION-001, REQ-REGION-002
Maps to erp-auto-test suite storage-regions (TC-STR-REG/RES, TC-REL-VIS, TC-UI-REL-VIS).
"""
from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from tcm_import_common import AUTHOR, ROLE_ADMIN, ROLE_OWNER, Case, Step

OUTPUT = (
    Path(__file__).resolve().parent.parent
    / "docs"
    / "tcm-import-visibility-regions-20260627.xlsx"
)

FEAT_ROOT = "REQ-REGION"
FEAT_LOC = "REQ-REGION-001"
FEAT_RES = "REQ-REGION-002"

PRE_ADMIN = "@Admin залогінений. Середовище dev/staging."
PRE_OWNER2 = (
    "@Owner 2 залогінений. Підрозділ owner2.storage.id з accessMode=REGIONS (RESTRICTED). "
    "У шапці SPA обрана локація Owner 2."
)
PRE_OWNER1 = "@Owner 1 залогінений. Підрозділ з accessMode=FULL_ACCESS."


def vis(
    test_id: str,
    feature_id: str,
    ac_id: str,
    title: str,
    description: str,
    *,
    priority: str = "HIGH",
    severity: str = "MAJOR",
    preconditions: str = PRE_ADMIN,
    expected_result: str = "",
    tags: str = "visibility-regions",
    role_name: str = ROLE_ADMIN,
    layer: str | None = "API",
    steps: list[tuple[str, str]],
    cross: list[str] | None = None,
) -> Case:
    return Case(
        test_id=test_id,
        feature_id=feature_id,
        ac_id=ac_id,
        title=title,
        description=description,
        priority=priority,
        severity=severity,
        preconditions=preconditions,
        expected_result=expected_result or (steps[-1][1] if steps else ""),
        tags=tags,
        role_name=role_name,
        automation_layer=layer,
        automation_test_id=test_id if layer else None,
        cross_features=cross or [],
        steps=[Step(i + 1, a, e) for i, (a, e) in enumerate(steps)],
    )


def location_region_cases() -> list[Case]:
    api = "visibility-regions,api"
    ui = "visibility-regions,ui"
    return [
        vis(
            "TC-STR-REG-001", FEAT_LOC, "AC-01",
            "Створення області видимості локацій",
            "POST /api/v1/storages/regions: name, recipientStorage, accessMode.",
            priority="CRITICAL", severity="CRITICAL", tags=api,
            steps=[
                ("Створити recipient-локацію", "storageId отримано"),
                ("POST /storages/regions з accessMode=FULL_ACCESS", "HTTP 200; id, name, recipientStorage у response"),
            ],
        ),
        vis(
            "TC-STR-REG-002", FEAT_LOC, "AC-02",
            "Список областей — фільтр по name",
            "GET /storages/regions?name= повертає створену область.",
            severity="MINOR", tags=api,
            steps=[
                ("Створити область з унікальним префіксом імені", "Область створена"),
                ("GET /storages/regions?name={prefix}", "HTTP 200; область у content"),
            ],
        ),
        vis(
            "TC-STR-REG-003", FEAT_LOC, "AC-03",
            "Деталі області за id",
            "GET /storages/regions/{id} — картка редагування.",
            priority="CRITICAL", severity="CRITICAL", tags=api,
            steps=[
                ("Створити область", "regionId отримано"),
                ("GET /storages/regions/{regionId}", "HTTP 200; id, name, recipientStorage.id збігаються"),
            ],
        ),
        vis(
            "TC-STR-REG-004", FEAT_LOC, "AC-04",
            "Оновлення області",
            "PUT змінює name, recipientStorage, accessMode.",
            priority="CRITICAL", severity="CRITICAL", tags=api,
            steps=[
                ("Створити область FULL_ACCESS", "regionId отримано"),
                ("PUT з новим name, іншим recipient, accessMode=REGIONS", "HTTP 200; поля оновлені"),
            ],
        ),
        vis(
            "TC-STR-REG-005", FEAT_LOC, "AC-05",
            "Видалення області",
            "DELETE області; GET by id не 200.",
            priority="CRITICAL", severity="CRITICAL", tags=api,
            steps=[
                ("Створити область", "regionId отримано"),
                ("DELETE /storages/regions/{regionId}", "HTTP 200"),
                ("GET /storages/regions/{regionId}", "HTTP ≠ 200"),
            ],
        ),
        vis(
            "TC-STR-REG-010", FEAT_LOC, "AC-06",
            "Додавання та видалення локацій у області",
            "PUT/DELETE /regions/{id}/locations.",
            priority="CRITICAL", severity="CRITICAL", tags=api,
            steps=[
                ("Створити область і локації A, B, C", "ids відомі"),
                ("PUT locations [A,B,C]; DELETE [B,C]", "GET locations містить лише A"),
            ],
        ),
        vis(
            "TC-STR-REG-012", FEAT_LOC, "AC-07",
            "Members області — підрозділи-споживачі",
            "PUT/DELETE /regions/{id}/members.",
            priority="CRITICAL", severity="CRITICAL", tags=api,
            steps=[
                ("Створити область; members A, B", "regionId відомий"),
                ("PUT members [A,B]; DELETE [B]", "GET members містить лише A"),
            ],
        ),
        vis(
            "TC-STR-REG-015", FEAT_LOC, "AC-08",
            "Autocomplete локацій для адмін-UI",
            "GET /storages/locations/suggest?name=.",
            severity="MINOR", tags=api,
            steps=[
                ("Створити STORAGE і прив'язати до області", "locationId у region"),
                ("GET /storages/locations/suggest?name={prefix}", "HTTP 200; локація у списку"),
            ],
        ),
        vis(
            "TC-STR-REG-RBAC-001", FEAT_LOC, "AC-09",
            "RBAC — Owner не керує областями видимості",
            "POST/PUT/DELETE /storages/regions під @Owner → 403.",
            priority="CRITICAL", preconditions=PRE_OWNER2, role_name=ROLE_OWNER,
            tags="visibility-regions,rbac,negative", layer=None,
            steps=[
                ("POST /storages/regions під Owner 2", "HTTP 403"),
                ("PUT /storages/regions/{id} під Owner 2", "HTTP 403"),
                ("DELETE /storages/regions/{id} під Owner 2", "HTTP 403"),
            ],
        ),
        vis(
            "TC-STR-REG-050", FEAT_LOC, "AC-10",
            "Explicit grant видимості локації",
            "PUT /storages/{visible}/locations?locations={viewer}.",
            priority="CRITICAL", severity="CRITICAL", tags=api,
            steps=[
                ("Створити visible-локацію та viewer-підрозділ", "ids відомі"),
                ("PUT explicit grant viewer←visible", "HTTP 200"),
                ("GET /storages/{visible}/locations", "viewer у списку links"),
            ],
        ),
        vis(
            "TC-STR-REG-050b", FEAT_LOC, "AC-11",
            "Revoke explicit grant",
            "DELETE explicit link; viewer більше не бачить локацію.",
            priority="CRITICAL", tags=api, layer=None,
            steps=[
                ("Надати explicit grant (setup)", "Grant активний"),
                ("DELETE /storages/{visible}/locations?locations={viewer}", "HTTP 200"),
                ("GET /storages/{visible}/locations", "viewer відсутній"),
            ],
        ),
        vis(
            "TC-STR-REG-054", FEAT_LOC, "AC-12",
            "Перегляд viewers для локації",
            "GET /storages/{visible}/locations — список підрозділів з grant.",
            severity="MINOR", tags=api,
            steps=[
                ("Grant + region member для однієї visible-локації", "Setup готовий"),
                ("GET /storages/{visible}/locations", "HTTP 200; viewers містить grant і member"),
            ],
        ),
        vis(
            "TC-STR-REG-020", FEAT_LOC, "AC-13",
            "Підрозділ accessMode=REGIONS (RESTRICTED)",
            "POST storage з accessMode=REGIONS.",
            priority="CRITICAL", severity="CRITICAL", tags=api,
            steps=[
                ("POST child storage accessMode=REGIONS", "HTTP 200; accessMode=REGIONS"),
            ],
        ),
        vis(
            "TC-STR-REG-022", FEAT_LOC, "AC-13",
            "RESTRICTED без областей — лише себе в /names",
            "GET /storages/names?isActive=true для Owner 2 без regions.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("Purge regions/grants для Owner 2", "Scope чистий"),
                ("GET /storages/names?isActive=true", "Рівно 1 запис — owner2.storage.id"),
            ],
        ),
        vis(
            "TC-STR-REG-021", FEAT_LOC, "AC-14",
            "Контраст FULL_ACCESS vs REGIONS у /names",
            "OWNER_1 ширший список; OWNER_2 — один id.",
            preconditions=PRE_OWNER1, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("GET /names під Owner 1", "Список > 1"),
                ("GET /names під Owner 2 (REGIONS, без областей)", "Рівно 1 id"),
            ],
        ),
        vis(
            "TC-STR-REG-033", FEAT_LOC, "AC-15",
            "/names/my-units для REGIONS owner",
            "Лише власний internal unit.",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, severity="MINOR", tags=api,
            steps=[
                ("GET /storages/names/my-units", "HTTP 200; рівно 1 unit — owner2"),
            ],
        ),
        vis(
            "TC-STR-REG-030", FEAT_LOC, "AC-16",
            "REGIONS-область — alias у /names",
            "Ім'я області замість recipient.name; outsider прихований.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("Region accessMode=REGIONS; member=Owner 2", "Setup готовий"),
                ("GET /storages/names?isActive=true", "region.name присутній; outsider відсутній"),
            ],
        ),
        vis(
            "TC-STR-REG-031", FEAT_LOC, "AC-17",
            "FULL_ACCESS-область — реальні імена локацій",
            "У /names — storage.name локацій з набору.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("Region FULL_ACCESS з locations; member=Owner 2", "Setup готовий"),
                ("GET /storages/names", "Реальні імена локацій з region.locations"),
            ],
        ),
        vis(
            "TC-STR-REG-040", FEAT_LOC, "AC-18",
            "Alias локації в FULL_ACCESS region",
            "name у /names = alias локації.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("STORAGE з alias; FULL_ACCESS region; member Owner 2", "Setup готовий"),
                ("GET /storages/names", "Запис з name=alias"),
            ],
        ),
        vis(
            "TC-STR-REG-034", FEAT_LOC, "AC-19",
            "Union трьох областей — без дублікатів id",
            "Спільна локація з 3 regions → один запис у /names.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("Owner 2 member у 3 областях з shared location", "Setup готовий"),
                ("GET /storages/names", "shared id з'являється рівно один раз"),
            ],
        ),
        vis(
            "TC-STR-REG-035", FEAT_LOC, "AC-19",
            "Union × типи STORAGE/UNIT/PRODUCTION та accessMode",
            "Параметризований контракт /names для різних типів.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("DataProvider: 3 regions, shared location типу T", "Setup для кожного типу"),
                ("GET /names", "Один запис без дублікатів; label за accessMode"),
            ],
        ),
        vis(
            "TC-STR-REG-014", FEAT_LOC, "AC-20",
            "Локація в двох областях",
            "GET locations обох regions містить shared storageId.",
            severity="MINOR", tags=api,
            steps=[
                ("sharedLocation у region1 і region2", "regions створені"),
                ("GET .../regions/{id1}/locations та .../{id2}/locations", "Обидва містять shared id"),
            ],
        ),
        vis(
            "TC-STR-REG-052", FEAT_LOC, "AC-21",
            "Explicit grant — реальне storage.name у /names",
            "Не alias області.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("Explicit grant visible←Owner 2", "Grant активний"),
                ("GET /storages/names", "name=visible.name (не region.name)"),
            ],
        ),
        vis(
            "TC-REL-VIS-001", FEAT_LOC, "AC-22",
            "Send в межах області видимості",
            "POST send in-scope → CREATED; stock −N.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation", "REQ-WMS-001"],
            steps=[
                ("Region member Owner 2; stock на sender", "Setup готовий"),
                ("POST /relocations/send на in-scope recipient", "HTTP 200; state=CREATED; stock −N"),
            ],
        ),
        vis(
            "TC-REL-VIS-007", FEAT_LOC, "AC-22",
            "Send на explicit grant локацію",
            "Grant дозволяє send.",
            priority="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation"],
            steps=[
                ("Explicit grant; stock Owner 2", "Setup готовий"),
                ("POST send на granted локацію", "HTTP 200; CREATED"),
            ],
        ),
        vis(
            "TC-REL-VIS-002", FEAT_LOC, "AC-23",
            "Send поза областю — заборонено",
            "Outsider не в /names → 4xx; stock без змін.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation"],
            steps=[
                ("Outsider storage поза regions Owner 2", "Setup готовий"),
                ("POST send recipient=outsider", "HTTP 4xx; stock sender не змінився"),
            ],
        ),
        vis(
            "TC-REL-VIS-009", FEAT_LOC, "AC-24",
            "Send на REGIONS alias → anchor recipient",
            "recipient.id = recipientStorage після FINISHED.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation"],
            steps=[
                ("Region REGIONS; send на alias з /names", "HTTP 200"),
                ("Resolve FINISHED", "Stock +N на anchor (recipientStorage), не на проміжну location"),
            ],
        ),
        vis(
            "TC-REL-VIS-008", FEAT_LOC, "AC-25",
            "Revoke grant — повторний send заборонено",
            "Після revoke explicit grant send → 4xx.",
            severity="MINOR",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation"],
            steps=[
                ("Grant → send OK → revoke grant", "Revoke виконано"),
                ("Повторний send на ту ж локацію", "HTTP 4xx"),
            ],
        ),
        vis(
            "TC-UI-REL-VIS-001", FEAT_LOC, "AC-26",
            "UI — outsider відсутній у «Кому відправляю»",
            "Dropdown = GET /names; outsider не показується.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, layer="UI", tags=ui,
            cross=["relocation", "REQ-WMS-001-01-02"],
            steps=[
                ("Відкрити форму видачі /relocations", "Форма завантажена"),
                ("Відкрити dropdown «Кому відправляю»", "In-scope присутній; outsider відсутній"),
            ],
        ),
        vis(
            "TC-UI-REL-VIS-002", FEAT_LOC, "AC-26",
            "UI — REGIONS alias у dropdown",
            "Опція = region.name, не recipient.name.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, layer="UI", tags=ui,
            cross=["relocation"],
            steps=[
                ("Region REGIONS; форма видачі", "Setup готовий"),
                ("Dropdown «Кому відправляю»", "Label = region.name"),
            ],
        ),
        vis(
            "TC-UI-REL-VIS-003", FEAT_LOC, "AC-24",
            "UI send на REGIONS alias → anchor",
            "POST send recipient.id = recipientStorage.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, layer="UI", tags=ui,
            cross=["relocation"],
            steps=[
                ("Обрати alias області у dropdown; заповнити список продукції", "Форма готова"),
                ("Підтвердити видачу", "HTTP 200; sent.recipient.id = anchor.id"),
            ],
        ),
        vis(
            "TC-UI-REL-010", FEAT_LOC, "AC-26",
            "UI dropdown без дублікатів (union regions)",
            "Combobox опції ⊆ API /names; shared location — один label.",
            priority="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, layer="UI", tags=ui,
            cross=["relocation"],
            steps=[
                ("GET /names API — підрахунок shared id", "1 entry"),
                ("UI dropdown labels", "Без дублікатів; опції ⊆ API names"),
            ],
        ),
        vis(
            "TC-STR-NAMES-001", FEAT_LOC, "AC-26",
            "Контракт GET /storages/names",
            "HTTP 200; schema; id + name для кожного елемента.",
            preconditions=PRE_OWNER1, role_name=ROLE_OWNER, severity="MINOR", tags=api,
            steps=[
                ("GET /storages/names?isActive=true", "HTTP 200; JSON schema; id+name не порожні"),
            ],
        ),
        vis(
            "TC-STR-NAMES-002", FEAT_LOC, "AC-19",
            "Унікальність id у /storages/names",
            "Без дублікатів для Combobox.",
            preconditions=PRE_OWNER1, role_name=ROLE_OWNER, severity="MINOR", tags=api,
            steps=[
                ("GET /storages/names?isActive=true", "size(unique ids) == size(list)"),
            ],
        ),
        vis(
            "TC-REL-VIS-003", FEAT_LOC, "AC-22",
            "Resolve FINISHED отримувачем у області",
            "Після in-scope send — resolve → stock +N.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation"],
            steps=[
                ("In-scope send CREATED", "Relocation створено"),
                ("POST resolve під Owner 2", "state=FINISHED; stock recipient +N"),
            ],
        ),
        vis(
            "TC-REL-VIS-004", FEAT_LOC, "AC-23",
            "Resolve з outsider storageId → 4xx",
            "Outsider не може resolve чужого переміщення.",
            priority="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation"],
            steps=[
                ("CREATED relocation in-scope", "id відомий"),
                ("Resolve storageId=outsider", "HTTP 4xx; state=CREATED"),
            ],
        ),
        vis(
            "TC-REL-VIS-005", FEAT_LOC, "AC-16",
            "REGIONS alias + send in-scope",
            "/names містить region.name; send дозволено.",
            priority="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation"],
            steps=[
                ("GET /names", "region.name присутній"),
                ("POST send in-scope", "HTTP 200 CREATED"),
            ],
        ),
    ]


def resource_region_cases() -> list[Case]:
    api = "visibility-resources,api"
    return [
        vis(
            "TC-STR-RES-001", FEAT_RES, "AC-01",
            "Область RESOURCES — CRUD ресурсів",
            "Створення region accessMode=RESOURCES; PUT/GET/DELETE .../resources.",
            priority="CRITICAL", severity="CRITICAL", tags=api,
            steps=[
                ("POST region accessMode=RESOURCES", "regionId отримано"),
                ("PUT resources [A,B]; GET list", "Обидва id у list"),
                ("DELETE resource B", "GET list містить лише A"),
            ],
        ),
        vis(
            "TC-STR-RES-RBAC-001", FEAT_RES, "AC-04",
            "RBAC — Owner не керує ресурсами області",
            "PUT/DELETE .../regions/{id}/resources під Owner → 403.",
            priority="CRITICAL", preconditions=PRE_OWNER2, role_name=ROLE_OWNER,
            tags="visibility-resources,rbac,negative", layer=None,
            steps=[
                ("PUT .../regions/{id}/resources під Owner 2", "HTTP 403"),
                ("DELETE .../regions/{id}/resources під Owner 2", "HTTP 403"),
            ],
        ),
        vis(
            "TC-STR-RES-002", FEAT_RES, "AC-05",
            "RESTRICTED без областей — autocomplete порожній",
            "GET /resources/autocomplete?storageId= без RESOURCES regions.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("Unit REGIONS без RESOURCES regions", "Setup готовий"),
                ("GET /resources/autocomplete?storageId={unit}", "Порожній або без тестових id"),
            ],
        ),
        vis(
            "TC-STR-RES-003", FEAT_RES, "AC-06",
            "Member бачить лише granted ресурси",
            "Autocomplete + GET page фільтрують hidden.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("Region RESOURCES: granted + hidden поза областю", "Setup готовий"),
                ("GET autocomplete та GET /resources page", "granted є; hidden відсутній"),
            ],
        ),
        vis(
            "TC-STR-RES-004", FEAT_RES, "AC-07",
            "Union двох областей RESOURCES",
            "Autocomplete = union без дублікатів id.",
            severity="MINOR",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("2 regions RESOURCES на одному unit; різні resources", "Setup готовий"),
                ("GET autocomplete", "Union обох наборів; unique ids"),
            ],
        ),
        vis(
            "TC-STR-RES-006", FEAT_RES, "AC-08",
            "FULL_ACCESS vs RESTRICTED — ширина номенклатури",
            "FULL autocomplete ширший за RESTRICTED+1 region.",
            severity="MINOR",
            preconditions=PRE_OWNER1, role_name=ROLE_OWNER, tags=api,
            steps=[
                ("FULL unit vs RESTRICTED з 1 granted", "Setup готовий"),
                ("Порівняти autocomplete count", "FULL > RESTRICTED"),
            ],
        ),
        vis(
            "TC-STR-RES-007", FEAT_RES, "AC-09",
            "Guard 2.1.1 — DELETE ресурсу з області при stock>0",
            "DELETE .../resources → 400 якщо stock>0 на unit.",
            severity="MINOR", tags=api,
            steps=[
                ("Resource у region; stock>0 через receive", "stock підтверджено"),
                ("DELETE resource з region", "HTTP 400; resource лишається в autocomplete"),
            ],
        ),
        vis(
            "TC-STR-RES-008", FEAT_RES, "AC-10",
            "Inventory PUT — ресурс поза областю → 400",
            "Guard 4.2: hidden resource у inventory session.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["REQ-WMS-003"],
            steps=[
                ("Inventory session на RESTRICTED unit; hidden поза region", "sessionId відомий"),
                ("PUT /inventory з hidden resource", "HTTP 400; stock hidden=0"),
            ],
        ),
        vis(
            "TC-STR-RES-010", FEAT_RES, "AC-11",
            "Inventory PUT — ресурс з області → 200",
            "Visible resource оновлює stock.",
            severity="MINOR",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["REQ-WMS-003"],
            steps=[
                ("Visible у RESOURCES region; inventory session", "sessionId відомий"),
                ("PUT /inventory з visible", "HTTP 200; stock оновлено"),
            ],
        ),
        vis(
            "TC-STR-RES-011", FEAT_RES, "AC-12",
            "Auto-grant 2.2 — INTERNAL receive outOfScope",
            "Receive розширює autocomplete + stock.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation", "REQ-WMS-001"],
            steps=[
                ("outOfScope поза region; send+resolve на RESTRICTED unit", "Receive виконано"),
                ("GET autocomplete", "outOfScope з'явився; stock>0"),
            ],
        ),
        vis(
            "TC-STR-RES-009", FEAT_RES, "AC-13",
            "SUPPLIER receive невидимого ресурсу (4.1 / 2.2)",
            "400 або auto-grant після receive.",
            severity="MINOR",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation"],
            steps=[
                ("Resource не в region; receive з SUPPLIER", "HTTP 400 або success"),
                ("GET autocomplete", "Resource у селекторі після auto-grant (2.2) або відсутній (4.1)"),
            ],
        ),
        vis(
            "TC-STR-RES-005", FEAT_RES, "AC-12",
            "Auto-grant після internal relocation receive",
            "Ресурс поза region з'являється в autocomplete.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation"],
            steps=[
                ("Resource поза region; receive internal на unit", "Receive FINISHED"),
                ("GET autocomplete storageId=unit", "Resource присутній"),
            ],
        ),
        vis(
            "TC-REL-VIS-010", FEAT_RES, "AC-14",
            "Send ресурсом з області RESOURCES",
            "POST send in-scope resource → 200.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation", "REQ-WMS-001-01-03"],
            steps=[
                ("Resource у RESOURCES region; stock на sender", "Setup готовий"),
                ("POST send з in-scope resource", "HTTP 200"),
            ],
        ),
        vis(
            "TC-REL-VIS-011", FEAT_RES, "AC-15",
            "Send ресурсом поза областю RESOURCES → 4xx",
            "Resource guard на send.",
            priority="CRITICAL", severity="CRITICAL",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, tags=api,
            cross=["relocation"],
            steps=[
                ("hidden resource не в autocomplete", "Setup готовий"),
                ("POST send з hidden resource (прямий API)", "HTTP 4xx"),
            ],
        ),
        vis(
            "TC-WMS-REG-RES-016", FEAT_RES, "AC-16",
            "UI — «Список продукції» лише in-scope ресурси",
            "RESTRICTED owner: autocomplete ресурсів на формі видачі фільтрується областями.",
            preconditions=PRE_OWNER2, role_name=ROLE_OWNER, layer="UI", tags="visibility-resources,ui",
            cross=["REQ-WMS-001-01-03"],
            steps=[
                ("Відкрити форму видачі під Owner 2 (REGIONS)", "Форма завантажена"),
                ("Autocomplete «Оберіть ресурс»", "Лише in-scope; hidden відсутній"),
            ],
        ),
    ]


def features() -> list[tuple]:
    return [
        (
            FEAT_ROOT, "",
            "Області видимості",
            "Централізоване керування видимістю локацій та ресурсів між підрозділами "
            "(accessMode REGIONS / FULL_ACCESS / RESOURCES, StorageRegion, explicit grants).",
            "WMS", "CRITICAL", "0", "2",
        ),
        (
            FEAT_LOC, FEAT_ROOT,
            "Області видимості локацій",
            "Named-набори локацій для RESTRICTED підрозділів: CRUD областей, members/locations, "
            "explicit grants, селектор GET /storages/names, інтеграція з переміщеннями.",
            "WMS", "CRITICAL", "1", "0",
        ),
        (
            FEAT_RES, FEAT_ROOT,
            "Області видимості ресурсів",
            "Області accessMode=RESOURCES: фільтр номенклатури для RESTRICTED підрозділів, "
            "guards inventory/send, auto-grant після receive.",
            "WMS", "CRITICAL", "1", "1",
        ),
    ]


def acceptance_criteria() -> list[tuple]:
    loc = [
        ("AC-01", "@Admin може створити область видимості: name, recipientStorage, accessMode (FULL_ACCESS / REGIONS)"),
        ("AC-02", "@Admin може переглянути список областей з фільтром по name"),
        ("AC-03", "@Admin може переглянути деталі області за id"),
        ("AC-04", "@Admin може оновити область (name, recipientStorage, accessMode)"),
        ("AC-05", "@Admin може видалити область"),
        ("AC-06", "@Admin може додати/видалити локації в області (.../locations)"),
        ("AC-07", "@Admin може додати/видалити підрозділи-споживачі області (.../members)"),
        ("AC-08", "@Admin може шукати локації для прив'язки через autocomplete (/locations/suggest)"),
        ("AC-09", "@Owner не може створювати/редагувати/видаляти області (лише @Admin)"),
        ("AC-10", "@Admin може надати explicit grant видимості локації підрозділу без області"),
        ("AC-11", "@Admin може відкликати explicit grant"),
        ("AC-12", "@Admin може переглянути viewers для локації (GET /storages/{id}/locations)"),
        ("AC-13", "Підрозділ accessMode=REGIONS без областей бачить у GET /storages/names лише себе"),
        ("AC-14", "Підрозділ accessMode=FULL_ACCESS бачить ширший набір локацій, ніж RESTRICTED без областей"),
        ("AC-15", "GET /storages/names/my-units для REGIONS-owner показує лише власний internal unit"),
        ("AC-16", "Область accessMode=REGIONS: у /names — ім'я області (alias), не recipient.name; outsider прихований"),
        ("AC-17", "Область accessMode=FULL_ACCESS: у /names — реальні імена локацій з набору"),
        ("AC-18", "Alias локації відображається в FULL_ACCESS-області"),
        ("AC-19", "Member у кількох областях отримує union видимості без дублікатів id"),
        ("AC-20", "Одна локація може входити в кілька областей"),
        ("AC-21", "Explicit grant показує реальне storage.name у /names (не alias області)"),
        ("AC-22", "@Owner може send на in-scope локацію (member області або explicit grant)"),
        ("AC-23", "@Owner не може send на локацію поза областями видимості"),
        ("AC-24", "Send на REGIONS-alias → recipient.id = recipientStorage (anchor)"),
        ("AC-25", "Після revoke explicit grant повторний send заборонено"),
        ("AC-26", "UI dropdown «Кому відправляю» = GET /names: без outsider; для REGIONS — ім'я області"),
    ]
    res = [
        ("AC-01", "@Admin може створити область з accessMode=RESOURCES"),
        ("AC-02", "@Admin може додати/видалити ресурси в області RESOURCES"),
        ("AC-03", "@Admin може переглянути список ресурсів області"),
        ("AC-04", "@Owner не може керувати ресурсами області"),
        ("AC-05", "RESTRICTED підрозділ без областей RESOURCES — autocomplete ресурсів порожній"),
        ("AC-06", "Member бачить лише granted-ресурси (autocomplete + GET page)"),
        ("AC-07", "Дві і більше областей RESOURCES для одного member → union ресурсів без дублікатів"),
        ("AC-08", "FULL_ACCESS підрозділ бачить ширшу номенклатуру, ніж RESTRICTED з однією областю"),
        ("AC-09", "@Admin не може видалити ресурс з області, якщо stock > 0 на підрозділі-споживачі (guard 2.1.1)"),
        ("AC-10", "@Owner не може inventory PUT з ресурсом поза областю RESOURCES (guard 4.2)"),
        ("AC-11", "@Owner може inventory PUT з ресурсом з області RESOURCES"),
        ("AC-12", "Internal receive outOfScope-ресурсу автоматично розширює видимість + stock (2.2)"),
        ("AC-13", "SUPPLIER receive невидимого ресурсу — поведінка за 4.1 (400) або auto-grant (2.2)"),
        ("AC-14", "Send ресурсом з області RESOURCES → дозволено"),
        ("AC-15", "Send ресурсом поза областю RESOURCES → 4xx"),
        ("AC-16", "Блок «Список продукції» на формі видачі показує лише in-scope ресурси для RESTRICTED"),
    ]
    rows: list[tuple] = []
    for i, (ac_id, text) in enumerate(loc):
        rows.append((FEAT_LOC, ac_id, text, str(i)))
    for i, (ac_id, text) in enumerate(res):
        rows.append((FEAT_RES, ac_id, text, str(i)))
    return rows


def all_cases() -> list[Case]:
    seen: set[str] = set()
    unique: list[Case] = []
    for case in location_region_cases() + resource_region_cases():
        if case.test_id in seen:
            continue
        seen.add(case.test_id)
        unique.append(case)
    return unique


def write_xlsx_with_features(cases: list[Case], path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    meta = wb.create_sheet("Meta")
    meta.append(["key", "value"])
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")
    for row in [
        ("formatVersion", "1"),
        ("exportedAt", now),
        ("projectId", "1"),
        ("projectName", "ERP Система"),
        ("scope", "PROJECT"),
        ("rootFeatureId", ""),
        ("source", "erp-auto-test visibility regions 2026-06-27"),
    ]:
        meta.append(list(row))

    feat_sheet = wb.create_sheet("Features")
    feat_sheet.append([
        "featureId", "parentFeatureId", "title", "description", "module",
        "priority", "status", "author", "treeDepth", "sortOrder",
    ])
    for row in features():
        feat_sheet.append([*row, "ACTIVE", AUTHOR])

    ac_sheet = wb.create_sheet("AcceptanceCriteria")
    ac_sheet.append(["featureId", "acId", "text", "sortOrder"])
    for row in acceptance_criteria():
        ac_sheet.append(list(row))

    tc_sheet = wb.create_sheet("TestCases")
    tc_sheet.append([
        "testId", "featureId", "acId", "title", "description", "priority", "severity", "status",
        "testType", "preconditions", "expectedResult", "tags", "author", "jiraIssueKey", "roleName",
        "parameterized", "dependencies",
    ])
    for case in cases:
        tc_sheet.append([
            case.test_id, case.feature_id, case.ac_id, case.title, case.description,
            case.priority, case.severity, "ACTIVE", case.test_type,
            case.preconditions, case.expected_result, case.tags, AUTHOR, "", case.role_name,
            "false", "",
        ])

    steps_sheet = wb.create_sheet("Steps")
    steps_sheet.append(["testId", "stepOrder", "actionText", "expectedText"])
    for case in cases:
        for step in case.steps:
            steps_sheet.append([case.test_id, str(step.order), step.action, step.expected])

    wb.create_sheet("DatasetSchema").append(
        ["testId", "fieldKey", "fieldLabel", "fieldType", "required", "sortOrder"]
    )
    wb.create_sheet("ParameterSets").append(["testId", "setName", "active", "valuesJson"])

    auto_sheet = wb.create_sheet("AutomationLinks")
    auto_sheet.append(["testId", "layer", "automationTestId", "sortOrder"])
    for case in cases:
        if case.automation_layer and case.automation_test_id:
            auto_sheet.append([case.test_id, case.automation_layer, case.automation_test_id, "0"])

    cross_sheet = wb.create_sheet("CrossFeatures")
    cross_sheet.append(["testId", "crossFeatureSlug"])
    for case in cases:
        for slug in case.cross_features:
            cross_sheet.append([case.test_id, slug])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    cases = all_cases()
    write_xlsx_with_features(cases, OUTPUT)
    ac_loc = sum(1 for r in acceptance_criteria() if r[0] == FEAT_LOC)
    ac_res = sum(1 for r in acceptance_criteria() if r[0] == FEAT_RES)
    auto = sum(1 for c in cases if c.automation_test_id)
    print(f"Wrote {OUTPUT}")
    print(f"  Features: 3")
    print(f"  AC: {ac_loc} (locations) + {ac_res} (resources) = {ac_loc + ac_res}")
    print(f"  TestCases: {len(cases)} ({auto} with AutomationLinks)")
    by_feat: dict[str, int] = {}
    for c in cases:
        by_feat[c.feature_id] = by_feat.get(c.feature_id, 0) + 1
    for k, v in sorted(by_feat.items()):
        print(f"    {k}: {v} tests")


if __name__ == "__main__":
    main()
