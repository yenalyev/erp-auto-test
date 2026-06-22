"""Shared TCM XLSX import helpers (TestCaseXlsxIO v1)."""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook

AUTHOR = "QA"
ROLE_OWNER = "Owner"
ROLE_ADMIN = "Admin"


@dataclass
class Step:
    order: int
    action: str
    expected: str


@dataclass
class Case:
    test_id: str
    feature_id: str
    ac_id: str
    title: str
    description: str
    priority: str = "HIGH"
    severity: str = "MAJOR"
    test_type: str = "FUNCTIONAL"
    preconditions: str = ""
    expected_result: str = ""
    tags: str = ""
    role_name: str = ROLE_OWNER
    steps: list[Step] = field(default_factory=list)
    automation_layer: str | None = None
    automation_test_id: str | None = None
    cross_features: list[str] = field(default_factory=list)


def write_xlsx(cases: list[Case], path: Path, project_name: str = "ERP Система") -> None:
    wb = Workbook()
    wb.remove(wb.active)

    meta = wb.create_sheet("Meta")
    meta.append(["key", "value"])
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")
    for row in [
        ("formatVersion", "1"),
        ("exportedAt", now),
        ("projectId", "1"),
        ("projectName", project_name),
        ("scope", "PROJECT"),
        ("rootFeatureId", ""),
    ]:
        meta.append(list(row))

    tc_sheet = wb.create_sheet("TestCases")
    tc_sheet.append([
        "testId", "featureId", "acId", "title", "description", "priority", "severity", "status",
        "testType", "preconditions", "expectedResult", "tags", "author", "jiraIssueKey", "roleName",
        "parameterized", "dependencies",
    ])
    for c in cases:
        tc_sheet.append([
            c.test_id, c.feature_id, c.ac_id, c.title, c.description,
            c.priority, c.severity, "ACTIVE", c.test_type,
            c.preconditions, c.expected_result, c.tags, AUTHOR, "", c.role_name,
            "false", "",
        ])

    steps_sheet = wb.create_sheet("Steps")
    steps_sheet.append(["testId", "stepOrder", "actionText", "expectedText"])
    for c in cases:
        for s in c.steps:
            steps_sheet.append([c.test_id, str(s.order), s.action, s.expected])

    schema_sheet = wb.create_sheet("DatasetSchema")
    schema_sheet.append(["testId", "fieldKey", "fieldLabel", "fieldType", "required", "sortOrder"])

    params_sheet = wb.create_sheet("ParameterSets")
    params_sheet.append(["testId", "setName", "active", "valuesJson"])

    auto_sheet = wb.create_sheet("AutomationLinks")
    auto_sheet.append(["testId", "layer", "automationTestId", "sortOrder"])
    for c in cases:
        if c.automation_layer and c.automation_test_id:
            auto_sheet.append([c.test_id, c.automation_layer, c.automation_test_id, "0"])

    cross_sheet = wb.create_sheet("CrossFeatures")
    cross_sheet.append(["testId", "crossFeatureSlug"])
    for c in cases:
        for slug in c.cross_features:
            cross_sheet.append([c.test_id, slug])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
